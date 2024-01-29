from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
import time
import pyautogui
import pandas as pd
import openpyxl
from datetime import datetime
import os
from datetime import date, timedelta 

os.startfile(R"Robo_robo.exe") 
     
file_path = 'Praticas.xlsx'
df = pd.read_excel(file_path)  

nome_da_coluna = 'PRÁTICA'
Pratica = df[nome_da_coluna].count()    
int_Pratica = int(Pratica)

nome_da_coluna = 'STATUS'
Status = df[nome_da_coluna].count()   

User = df.iloc[0, 4]
User_x = int(User)
                     
Senha = df.iloc[0, 5]
Senha_x = str(Senha)

User = "509544"

servico = Service(ChromeDriverManager().install())
nav = webdriver.Chrome(service=servico)
nav.maximize_window()
time.sleep(1.0)

                
nav.get('https://nsprd.sabesp.com.br/netainf/login.aspx')#URL NET@
time.sleep(1.0)
nav.find_element(By.ID,'extended_login_Username').send_keys(User_x)#USER
time.sleep(0.5)
nav.find_element(By.ID,'extended_login_Password').send_keys(Senha_x)#SENHA
time.sleep(0.5)
nav.find_element(By.ID,'extended_login_Login').click()#BOATO LOGIN
time.sleep(0.5)
nav.find_element(By.ID,'365').click()#CRM
time.sleep(0.5)  
nav.find_element(By.XPATH,'//*[@id="POR_153_10__365"]/div[2]').click()#PESQUISA PROCESSOS
time.sleep(0.5)

for i in range(Status, len(df)):

    valor_celula = df.iloc[i, 0]
    valor_Res = df.iloc[i, 1]
    print(f"{i + 1} - {valor_celula}")
    x = (valor_celula)
    X_int = str(x)
    ff = "1"
    data_x = "2023"
    nav.find_element(By.ID,'ctl00_NetSiuCPH_txt_crm_riproc_annocodiceda').send_keys(ff)#Código Único
    time.sleep(1.5)
    pyautogui.press('backspace', presses=10)
    time.sleep(1.5)
    nav.find_element(By.ID,'ctl00_NetSiuCPH_txt_crm_riproc_annocodiceda').send_keys(data_x)#Código Único
    time.sleep(1.5)
    nav.find_element(By.ID,'ctl00_NetSiuCPH_txt_crm_riproc_ticket').send_keys(X_int)#Código Único
    time.sleep(1.5)
    nav.find_element(By.ID,'ctl00_cphRight_btn_cercaelenco').click()#PESQUISA
    time.sleep(4.0)
    nav.find_element(By.ID,'ctl00_NetSiuCPH_gv_elencoproc_ctl02_cbx_crm_bpl_seleziona').click()#SELECIONAR
    time.sleep(1.5)
    nav.find_element(By.ID,'ctl00_cphRight_btnAvanzamentoMassivo').click()#EVOLUIR
    time.sleep(8.0)

    img = pyautogui.locateOnScreen('resources\Erro.png')
    img_2 = pyautogui.locateOnScreen('resources\Info.png')

    if img:

        time.sleep(1.5)
        nav.find_element(By.ID,'ctl00_nmp_Avvisi_Errore_btnYes').click()#OK

        def encontrar_proxima_linha_vazia_na_coluna(sheet, coluna):
            linha = 1
            while sheet.cell(row=linha, column=coluna).value is not None:
                linha += 1
            return linha

        try:
            wb = openpyxl.load_workbook('Praticas.xlsx')
        except FileNotFoundError:
            wb = openpyxl.Workbook()

        if 'Sheet' in wb.sheetnames:
                ws = wb['Sheet']
        else:
            ws = wb.active

        Hoje = datetime.now()
        Data = Hoje.strftime("%d/%m/%Y")
        Hora = Hoje.strftime("%H:%M:%S")
        
        proxima_linha_b = encontrar_proxima_linha_vazia_na_coluna(ws, coluna=2)
        ws.cell(row=proxima_linha_b, column=2, value="OS Ja Fechada")

        proxima_linha_c = encontrar_proxima_linha_vazia_na_coluna(ws, coluna=3)
        ws.cell(row=proxima_linha_c, column=3, value= Data)

        proxima_linha_d = encontrar_proxima_linha_vazia_na_coluna(ws, coluna=4)
        ws.cell(row=proxima_linha_d, column=4, value= Hora)
                    
        wb.save('Praticas.xlsx')

    else:

        if img_2:
            time.sleep(1.5)
            nav.find_element(By.XPATH,'//*[@id="ctl00_nmp_Avvisi_Errore_btnYes"]').click()#OK

            def encontrar_proxima_linha_vazia_na_coluna(sheet, coluna):
                linha = 1
                while sheet.cell(row=linha, column=coluna).value is not None:
                    linha += 1
                return linha

            try:
                wb = openpyxl.load_workbook('Praticas.xlsx')
            except FileNotFoundError:
                wb = openpyxl.Workbook()

            if 'Sheet' in wb.sheetnames:
                    ws = wb['Sheet']
            else:
                ws = wb.active

            Hoje = datetime.now()
            Data = Hoje.strftime("%d/%m/%Y")
            Hora = Hoje.strftime("%H:%M:%S")
            
            proxima_linha_b = encontrar_proxima_linha_vazia_na_coluna(ws, coluna=2)
            ws.cell(row=proxima_linha_b, column=2, value="Cancelamento em curso")

            proxima_linha_c = encontrar_proxima_linha_vazia_na_coluna(ws, coluna=3)
            ws.cell(row=proxima_linha_c, column=3, value= Data)

            proxima_linha_d = encontrar_proxima_linha_vazia_na_coluna(ws, coluna=4)
            ws.cell(row=proxima_linha_d, column=4, value= Hora)
                        
            wb.save('Praticas.xlsx')

        else: 
            time.sleep(2.0)
            nav.switch_to.frame('NETAModalDialogiFrame_1')
            element_inside_iframe = WebDriverWait(nav, 10).until(
                EC.presence_of_element_located((By.XPATH, '//*[@id="ctl00_NetSiuCPH_ddl_comando_avanzamento"]'))#Seta
            )
            nav.find_element(By.ID,'ctl00_NetSiuCPH_ddl_comando_avanzamento').click()#Confirmar
            time.sleep(1.0)
            pyautogui.press('down')
            time.sleep(1.0)
            pyautogui.press('enter')
            time.sleep(2.0)
            nav.find_element(By.ID,'ctl00_cphRight_btn_conferma').click()#Confirmar
            nav.switch_to.default_content()
            time.sleep(3.0)
            nav.find_element(By.ID,'ctl00_nmp_Avvisi_Errore_btnYes').click()#ok

            def encontrar_proxima_linha_vazia_na_coluna(sheet, coluna):
                linha = 1
                while sheet.cell(row=linha, column=coluna).value is not None:
                    linha += 1
                return linha

            try:
                wb = openpyxl.load_workbook('Praticas.xlsx')
            except FileNotFoundError:
                wb = openpyxl.Workbook()

            if 'Sheet' in wb.sheetnames:
                ws = wb['Sheet']
            else:
                ws = wb.active

            Hoje = datetime.now()
            Data = Hoje.strftime("%d/%m/%Y")
            Hora = Hoje.strftime("%H:%M:%S")
        
            proxima_linha_b = encontrar_proxima_linha_vazia_na_coluna(ws, coluna=2)
            ws.cell(row=proxima_linha_b, column=2, value="OS Fechada")

            proxima_linha_c = encontrar_proxima_linha_vazia_na_coluna(ws, coluna=3)
            ws.cell(row=proxima_linha_c, column=3, value= Data)

            proxima_linha_d = encontrar_proxima_linha_vazia_na_coluna(ws, coluna=4)
            ws.cell(row=proxima_linha_d, column=4, value= Hora)

            wb.save('Praticas.xlsx')

    time.sleep(2.0)
    nav.find_element(By.ID,'ctl00_NetSiuCPH_btn_visfiltri').click()#BOTAO PESQUISA   
    time.sleep(1.5) 
    nav.find_element(By.ID,'ctl00_cphRight_btn_reset').click()#Limpar
    time.sleep(1.5)   