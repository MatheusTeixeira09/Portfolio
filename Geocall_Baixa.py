from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
import time
import pyautogui
import pandas as pd
import openpyxl
from datetime import datetime
import os
from datetime import date, timedelta 
os.startfile(R"Robo_Robo_Geocall.exe")

#Data frame Cortes
file_path = 'Cortes.xlsx'
df = pd.read_excel(file_path) 

#Coluna Número OS - Cortes
nome_da_coluna = 'Número OS'
Pratica = df[nome_da_coluna].count()    
int_Pratica = int(Pratica)

#Coluna STATUS - BAIXAS - Cortes
nome_da_coluna = 'STATUS - BAIXA'
Status = df[nome_da_coluna].count()

#data
Data = df.iloc[0, 6]
Data_certa = Data.strftime("%d%m%Y")
print(Data_certa)

#Data frame Login
file_path_1 = 'Login_Geocall.xlsx'
df_1 = pd.read_excel(file_path_1)

#Escolha tipo de Baixa
Tipo_de_baixa = df.iloc[0, 7]

Motivo = df.iloc[0, 9]

#Escolha contrato
contrato = df.iloc[0, 8]
if contrato == 'Norte':

    User = df_1.iloc[0, 0]
    User_x = int(User)
                        
    Senha = df_1.iloc[0, 1]
    Senha_x = str(Senha)

else:

    User = df_1.iloc[1, 0]
    User_x = int(User)
                        
    Senha = df_1.iloc[1, 1]
    Senha_x = str(Senha)

#Matricula
Matricula = df.iloc[0, 5]
Matricula_x = str(Matricula)

#Declarar webdriver
servico = Service(ChromeDriverManager().install())
nav = webdriver.Chrome(service=servico)
nav.maximize_window()
time.sleep(1.0)

#URL NET@
nav.get('https://geoprd.sabesp.com.br/sabespwfm/')
time.sleep(1.0)

#Declarar frame
nav.switch_to.frame('mainFrame')
iframe_1 = WebDriverWait(nav, 10).until(
    EC.presence_of_element_located((By.XPATH, '//*[@id="USER"]'))
)

#USER
nav.find_element(By.ID,'USER').send_keys(User_x)
time.sleep(0.5)

#Senha
nav.find_element(By.ID,'INPUTPASS').send_keys(Senha_x)
time.sleep(0.5)

#Entrar
nav.find_element(By.ID,'submbtn').click()
time.sleep(0.5)

#Planejamento
nav.find_element(By.XPATH,'//*[@id="TBB_tbm2"]/div[4]').click()
time.sleep(0.5)

#Inicio da repetição 
for i in range(Status, len(df)):

    #Pegar o primeiro número os 
    valor_celula = df.iloc[i, 0]
    valor_Res = df.iloc[i, 1]
    print(f"{i + 1} - {valor_celula}")
    x = (valor_celula)
    X_int = str(x)

    #Busca Execução
    nav.find_element(By.XPATH,'//*[@id="TBB_tbm2"]/div[5]').click()
    time.sleep(1.5)

    #Abrir Recurso
    Recurso_Norte = pyautogui.locateCenterOnScreen('Imag_Baixa_geocall\Recurso.PNG')
    if Recurso_Norte:
        print("1")
        pyautogui.click('Imag_Baixa_geocall\Recurso.PNG')
        time.sleep(1.)
        pyautogui.click('Imag_Baixa_geocall\Editar_Recurso.PNG')
        time.sleep(1.5)
        pyautogui.write(Matricula_x)
        time.sleep(0.5)
        pyautogui.press('backspace', presses=2)
        time.sleep(0.5)
        pyautogui.press('enter')
        time.sleep(1.0)
    else:
        time.sleep(0.5)

    #Click no nome
    Nome_Norte = pyautogui.locateCenterOnScreen('Imag_Baixa_geocall\Y_nome_norte.PNG')
    Nome_Centro = pyautogui.locateCenterOnScreen('Imag_Baixa_geocall\Y_Nome_centro.PNG')
    if Nome_Norte:
        time.sleep(0.5)
        pyautogui.click('Imag_Baixa_geocall\Y_nome_norte.PNG')
        time.sleep(0.5)

    elif Nome_Centro:
        time.sleep(0.5)
        pyautogui.click('Imag_Baixa_geocall\Y_Nome_centro.PNG')
        time.sleep(0.5)
    else:
        pyautogui.click(x=510, y=520)
        time.sleep(1.0)
    
    
    #Selecionar
    Selecionar_Norte = pyautogui.locateOnScreen('Imag_Baixa_geocall\Selecionar_Norte.PNG')
    Selecionar_Centro = pyautogui.locateOnScreen('Imag_Baixa_geocall\Selecionar_Centro.PNG')
    if Selecionar_Norte:

        pyautogui.click('Imag_Baixa_geocall\Selecionar_Norte.PNG')

    elif Selecionar_Centro:

        pyautogui.click('Imag_Baixa_geocall\Selecionar_Centro.PNG')

    else:

        pyautogui.click(x=600, y=825)
        

    time.sleep(1.0)

    #Abrir Dados
    Dados_Norte = pyautogui.locateCenterOnScreen('Imag_Baixa_geocall\Dados_OS.png')
    if Dados_Norte:
        pyautogui.click('Imag_Baixa_geocall\Dados_OS.png')
    else:
        pyautogui.click('Imag_Baixa_geocall\Dados_OS_Centro.png')
    time.sleep(0.5)

    #Numero OsS
    Os_Norte = pyautogui.locateCenterOnScreen('Imag_Baixa_geocall\OS.png')
    if Os_Norte:
        pyautogui.click('Imag_Baixa_geocall\OS.png')
    else:
        pyautogui.click('Imag_Baixa_geocall\OS_Centro.png')
    time.sleep(0.5)
    pyautogui.press('tab') 
    time.sleep(0.5)
    pyautogui.write(X_int) 
    time.sleep(0.5)
    pyautogui.press('Enter') 
    time.sleep(1.5)

    Menu = pyautogui.locateOnScreen('Imag_Baixa_geocall\Menu.png')
    Menu_2 = pyautogui.locateOnScreen('Imag_Baixa_geocall\Menu_2.png')

    if Menu or Menu_2:
        print('Achou')
        time.sleep(1.5)
        nav.find_element(By.XPATH,'//*[@id="TV-stvRicerca"]/div/div/div/table/tbody/tr[2]/td[1]/div[1]/div/div/img').click()#Menu
        time.sleep(1.0
                   )
        pyautogui.click('Imag_Baixa_geocall\Resultado.png')#Resultado
        time.sleep(1.5)
        Validada = pyautogui.locateOnScreen('Imag_Baixa_geocall\Validada.png')
        Aberta = pyautogui.locateOnScreen('Imag_Baixa_geocall\Aberta.png')

        time.sleep(1.5)

        if Validada or Aberta:
            Hoje = datetime.now()
            Data = Hoje.strftime("%d/%m/%Y")
            Hora = Hoje.strftime("%H:%M:%S")
            time.sleep(1.5)
            pyautogui.scroll(1600)
            pyautogui.click('Imag_Baixa_geocall\Data.png')
            time.sleep(1.5)
            pyautogui.press('tab') 
            time.sleep(0.5)
            pyautogui.write(Data) 
            pyautogui.press('tab') 
            time.sleep(1.5)
            pyautogui.press('tab') 
            time.sleep(0.5)
            pyautogui.press('tab') 
            time.sleep(0.5)
            pyautogui.write(Hora) 
            time.sleep(0.5)
            pyautogui.press('tab') 
            time.sleep(0.5)
            pyautogui.press('tab')
            time.sleep(0.5)
            pyautogui.write(Data) 
            time.sleep(0.5)
            pyautogui.press('tab') 
            time.sleep(0.5)
            pyautogui.press('tab')
            time.sleep(0.5)
            pyautogui.write(Hora)
            time.sleep(0.5)
            pyautogui.press('tab') 
            time.sleep(0.5)
            pyautogui.press('tab')
            time.sleep(0.5)
            pyautogui.write(Data) 
            time.sleep(0.5)
            pyautogui.press('tab') 
            time.sleep(0.5)
            pyautogui.press('tab')
            time.sleep(0.5)
            pyautogui.write(Hora)
            time.sleep(0.5)
            pyautogui.press('tab') 
            time.sleep(0.5)
            pyautogui.press('tab')
            time.sleep(0.5)
            pyautogui.write(Data) 
            time.sleep(0.5)
            pyautogui.press('tab') 
            time.sleep(0.5)
            pyautogui.press('tab')
            time.sleep(0.5)
            pyautogui.write(Hora)
            time.sleep(1.5)
            pyautogui.click('Imag_Baixa_geocall\Causa.png')
            time.sleep(1.5)

            if Tipo_de_baixa == "CLIENTE APRESENTOU AS FATURAS PAGAS":
                pyautogui.click('Imag_Baixa_geocall\Conta_paga.png')
                time.sleep(1.5)
                print("CLIENTE APRESENTOU AS FATURAS PAGAS")

            elif Tipo_de_baixa == "IMÓVEL FECHADO, ORDEM ENCERRADA":
                pyautogui.click('Imag_Baixa_geocall\Imovel_fechado.png')
                time.sleep(1.5)
                print("IMÓVEL FECHADO, ORDEM ENCERRADA")
            
            time.sleep(5.0)          
            pyautogui.press('tab', presses=17)
            time.sleep(1.0)
            pyautogui.write(Motivo)
            time.sleep(1.0)  
            pyautogui.scroll(-1600)
            time.sleep(1.0)
            pyautogui.moveTo(100, 200) 
            time.sleep(1.5)
            pyautogui.click('Imag_Baixa_geocall\Salvar.png')
            time.sleep(10.0)
            Fechar_x = pyautogui.locateOnScreen('Imag_Baixa_geocall\Fechar_x.png')

            if Fechar_x:
                pyautogui.click('Imag_Baixa_geocall\Fechar_x.png')
                time.sleep(1.5)
                pyautogui.click('Imag_Baixa_geocall\Validar_OS.PNG')
                time.sleep(2.0)
                pyautogui.press('Enter') 
                time.sleep(1.5)
                pyautogui.click('Imag_Baixa_geocall\Fechar.png')
                time.sleep(2.0)

            else:
                pyautogui.click('Imag_Baixa_geocall\Fechar.png')
                time.sleep(2.0)

            
            def encontrar_proxima_linha_vazia_na_coluna(sheet, coluna):
                linha = 1
                while sheet.cell(row=linha, column=coluna).value is not None:
                    linha += 1
                return linha

            try:
                wb = openpyxl.load_workbook('Cortes.xlsx')
            except FileNotFoundError:
                wb = openpyxl.Workbook()

            if 'Sheet' in wb.sheetnames:
                    ws = wb['Sheet']
            else:
                ws = wb.active
            
            proxima_linha_b = encontrar_proxima_linha_vazia_na_coluna(ws, coluna=5)
            ws.cell(row=proxima_linha_b, column=5, value="OS Fechada")
                        
            wb.save('Cortes.xlsx')

            time.sleep(2.0)
        
        else:
            print('NAO ACHOU VALIDADA OU ABERTA')
            time.sleep(5.0)
            pyautogui.click('Imag_Baixa_geocall\Fechar.png')
            def encontrar_proxima_linha_vazia_na_coluna(sheet, coluna):
                linha = 1
                while sheet.cell(row=linha, column=coluna).value is not None:
                    linha += 1
                return linha

            try:
                wb = openpyxl.load_workbook('Cortes.xlsx')
            except FileNotFoundError:
                wb = openpyxl.Workbook()

            if 'Sheet' in wb.sheetnames:
                    ws = wb['Sheet']
            else:
                ws = wb.active
            
            proxima_linha_b = encontrar_proxima_linha_vazia_na_coluna(ws, coluna=5)
            ws.cell(row=proxima_linha_b, column=5, value="OS Ja Fechada")
                        
            wb.save('Cortes.xlsx')

            time.sleep(2.0)
        

    else:
        print('NAO ACHOU')
        time.sleep(2.0)
        def encontrar_proxima_linha_vazia_na_coluna(sheet, coluna):
            linha = 1
            while sheet.cell(row=linha, column=coluna).value is not None:
                linha += 1
            return linha

        try:
            wb = openpyxl.load_workbook('Cortes.xlsx')
        except FileNotFoundError:
            wb = openpyxl.Workbook()

        if 'Sheet' in wb.sheetnames:
                ws = wb['Sheet']
        else:
            ws = wb.active
        
        proxima_linha_b = encontrar_proxima_linha_vazia_na_coluna(ws, coluna=5)
        ws.cell(row=proxima_linha_b, column=5, value="OS Nao Encontrada")

                    
        wb.save('Cortes.xlsx')

        time.sleep(2.0)
print("acabou")