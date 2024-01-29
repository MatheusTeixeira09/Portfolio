from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
import time
import pyautogui
import pandas as pd
import openpyxl
from datetime import datetime
import os
from datetime import date, timedelta 

os.startfile(R"Robo_Robo_consulta.exe")

file_path_1 = 'Settings.xlsx'
df = pd.read_excel(file_path_1)  

User = df.iloc[0, 0]
User_x = int(User)

Senha = df.iloc[0, 1]
Senha_x = str(Senha)


file_path = 'Consulta.xlsx'
df_1 = pd.read_excel(file_path)  

nome_da_coluna = 'FORNECIMENTO'
Pratica = df_1[nome_da_coluna].count()    
int_Pratica = int(Pratica)

nome_da_coluna = 'STATUS NETA'
Status = df_1[nome_da_coluna].count()


servico = Service(ChromeDriverManager().install())
nav = webdriver.Chrome(service=servico)
nav.maximize_window()
time.sleep(1.0)

nav.get('https://nsprd.sabesp.com.br/netainf/login.aspx')#URL NET@
time.sleep(1.0)

nav.find_element(By.ID,'extended_login_Username').send_keys(User_x)#USER
time.sleep(0.5)

nav.find_element(By.ID,'extended_login_Password').send_keys(Senha_x)#SENHA
time.sleep(0.5)

nav.find_element(By.ID,'extended_login_Login').click()#BOATO LOGIN
time.sleep(1.5)


try:
    LOGIN = WebDriverWait(nav, 2).until(
        EC.presence_of_element_located((By.ID, 'valSum_extended'))
    )
    print("Erro")
    pyautogui.alert(text='Senha ou usuário errado', title='Erro', button='OK')
    nav.quit()
    time.sleep(0.5)
    os.startfile(R"resources\taskkill.bat")  


except TimeoutException:
    print("Senha certa")

    time.sleep(2.0)

    nav.get('https://nsprd.sabesp.com.br/NETASIU/SIUWeb/SiuWeb.Crm/Forms/DashBoards/RicercaCliente.aspx')#URL NET@
    time.sleep(2.0)

    

    for i in range(Status, len(df_1)):

        valor_celula = df_1.iloc[i, 0]
        valor_Res = df_1.iloc[i, 1]
        print(f"{i + 1} - {valor_celula}")
        x = (valor_celula)
        X_int = str(x)

        nav.find_element(By.ID,'ctl00_NetSiuCPH_txt_crm_rc_v_utenza').send_keys(X_int)#Fornecimento
        time.sleep(1.0)
        nav.find_element(By.ID,'ctl00_NetSiuCPH_btn_crm_rc_dettaglio').click()# BUSCA DETALHE
        time.sleep(1.5)

        n = 1
        while (n <= 1000):

            elemento = nav.find_element(By.ID,'ctl00_upProgress')

            if elemento.is_displayed():
                print("O elemento está visível na página.")

            else:
                print("O elemento não está visível na página.")
                break
                
            n += 1 

        #1° frame - para Status Fornecimento 
        nav.switch_to.frame('ifCruscottoUtenza')
        iframe_1 = WebDriverWait(nav, 10).until(
            EC.presence_of_element_located((By.ID,'ctl00_NetSiuCPH_lbl_crm_cruf_valstatofor'))#Status Fornecimento 
        )

        x = nav.find_element(By.ID, 'ctl00_NetSiuCPH_lbl_crm_cruf_valstatofor')#Status Fornecimento 
        texto = x.text
        print(texto)

        def verificar_palavras(texto, palavras_alvo):
                for palavra_alvo in palavras_alvo:
                    if palavra_alvo.lower() in texto.lower():
                        print(f'A palavra "{palavra_alvo}" está presente no texto.')
                        return True
                print('Nenhuma das palavras-alvo foi encontrada no texto.')
                return False

        palavras_alvo = ["ATIVO", "CORTADO", "SUPRIMIDO"]
        condicao = verificar_palavras(texto, palavras_alvo)

        if condicao:
            print('A condição é verdadeira.')

            time.sleep(5.0)
            nav.find_element(By.ID,'ctl00_NetSiuCPH_rep_crm_desktop_utenza_ctl01_lb_crm_cruf_df_zoom').click()#DOCUMENTOS
            time.sleep(6.0)
            nav.switch_to.default_content()

            #1° frame - para Status dovumentod
            nav.switch_to.frame('NETAModalDialogiFrame_1')
            iframe_2 = WebDriverWait(nav, 10).until(
                EC.presence_of_element_located((By.ID,'ctl00_NetSiuCPH_ctl01_ddl_crm_bdoc_filtro'))#DOCUMENTOS FILTRO 
            )

            nav.find_element(By.ID,'ctl00_NetSiuCPH_ctl01_ddl_crm_bdoc_filtro').click()#DOCUMENTOS FILTRO
            time.sleep(1.5)
            nav.find_element(By.XPATH,'//*[@id="ctl00_NetSiuCPH_ctl01_ddl_crm_bdoc_filtro"]/option[2]').click()#AFATURAS NAO PAGAS
            time.sleep(1.5)

            n = 1
            while (n <= 1000):

                elemento = nav.find_element(By.ID,'ctl00_upProgress')

                if elemento.is_displayed():
                    print("O elemento está visível na página.")

                else:
                    print("O elemento não está visível na página.")
                    break
                
                n += 1 

            time.sleep(1.0)

            #teste
        
            palavra1 = "OS"
            palavra2 = "OC"
            palavra3 = "D"
            palavra4 = "C"
            palavra5 = "J"
            palavra6 = "U"        

            OS = 0
            OS_ERRO = 0
            OS_CONT = 0

            OC = 0
            OC_ERRO = 0
            OC_CONT = 0

            D = 0
            D_ERRO = 0
            D_CONT = 0

            C = 0
            C_ERRO = 0
            C_CONT = 0

            CJ = 0
            CJ_ERRO = 0
            CJ_CONT = 0

            COBJUD = 0
            COBJUD_ERRO = 0
            COBJUD_CONT = 0
            
            try:
                tabela = WebDriverWait(nav, 10).until(
                    EC.presence_of_element_located((By.ID, "ctl00_NetSiuCPH_ctl01_ngv_crm_bdoc"))
                )
                indice_coluna_desejada = 10
                pelo_menos_uma_palavra_encontrada = False

                linhas = tabela.find_elements(By.TAG_NAME,"tr")
                for linha in linhas:
                    # Para cada linha, extraia as células (td)
                    celulas = linha.find_elements(By.TAG_NAME, "td")  
                    
                    if len(celulas) > indice_coluna_desejada:
                        celula_desejada = celulas[indice_coluna_desejada]
                        
                        print(f"Conteúdo da coluna {indice_coluna_desejada + 1}: {celula_desejada.text} ")

                        if palavra1.lower() in celula_desejada.text.lower():
                            print(f"Pelo menos uma das palavras '{palavra1}' foi encontrada na Célula {indice_coluna_desejada + 1} da linha.")
                            #pelo_menos_uma_palavra_encontrada = True 

                            indice_outra_coluna = 9

                            if len(celulas) > indice_outra_coluna:
                                outra_celula = celulas[indice_outra_coluna]
                                OS_CONT += 1
                                print(f"Conteúdo da outra coluna {indice_outra_coluna + 1}: {outra_celula.text}: {OS_CONT}")

                                if outra_celula.text == "N":
                                    print("Deu certo")

                                    OS += 1
                                    print(OS)
                                else:
                                    OS_ERRO += 1
                                    print(OS_ERRO)            
                                            
                            if OS > 0 :#OS

                                def encontrar_proxima_linha_vazia_na_coluna(sheet, coluna):
                                    linha = 1
                                    while sheet.cell(row=linha, column=coluna).value is not None:
                                        linha += 1
                                    return linha

                                try:
                                    wb = openpyxl.load_workbook('Consulta.xlsx')
                                except FileNotFoundError:
                                    wb = openpyxl.Workbook()

                                if 'Sheet' in wb.sheetnames:
                                    ws = wb['Sheet']
                                else:
                                    ws = wb.active

                                Hoje = datetime.now()
                                Data = Hoje.strftime("%d/%m/%Y")
                                Hora = Hoje.strftime("%H:%M:%S")    

                                proxima_linha_b = encontrar_proxima_linha_vazia_na_coluna(ws, coluna=4)
                                ws.cell(row=proxima_linha_b, column=4, value= texto )

                                proxima_linha_C = encontrar_proxima_linha_vazia_na_coluna(ws, coluna=5)
                                ws.cell(row=proxima_linha_C, column=5, value= "DEBITO EM ABERTO - ORDEM DE SUPRESSAO" )

                                proxima_linha_D = encontrar_proxima_linha_vazia_na_coluna(ws, coluna=6)
                                ws.cell(row=proxima_linha_D, column=6, value= Data )

                                proxima_linha_E = encontrar_proxima_linha_vazia_na_coluna(ws, coluna=7)
                                ws.cell(row=proxima_linha_E, column=7, value= Hora )
                                                
                                wb.save('Consulta.xlsx')

                                nav.switch_to.default_content()
                                nav.find_element(By.XPATH,'//*[@id="NETAModalDialog"]/div[1]/div[1]/button').click()#fECHARs
                                time.sleep(1.5)
                                nav.find_element(By.XPATH,'//*[@id="ctl00_NetSiuCPH_TabCRM_bli_tcrm_cruscotti"]/li/a/span/input').click()#FECHAR
                                time.sleep(1.0)

                                break

                            elif OS_ERRO == OS_CONT:

                                palavra7 = "FATURA ACORDO"
                                indice_coluna_desejada_fatura = 2

                                pelo_menos_uma_palavra_encontrada_X = False
                                pelo_menos_uma_palavra_encontrada_Y = False

                                linhas = tabela.find_elements(By.TAG_NAME,"tr")
                                for linha in linhas:
                                    # Para cada linha, extraia as células (td)
                                    celulas = linha.find_elements(By.TAG_NAME, "td") 


                                    if len(celulas) > indice_coluna_desejada_fatura:
                                        celula_desejada_fatura = celulas[indice_coluna_desejada_fatura]
                                        print(f"Conteúdo da coluna {indice_coluna_desejada_fatura + 1}: {celula_desejada_fatura.text}")

                                        if palavra7.lower() in celula_desejada_fatura.text.lower() :
                                            print(f"Pelo menos uma das palavras '{palavra7}' foi encontrada na Célula {indice_coluna_desejada_fatura + 1} da linha.")
                                            pelo_menos_uma_palavra_encontrada_X = True
                                    
                                if pelo_menos_uma_palavra_encontrada_X:#FATURA ACORDO

                                    nav.switch_to.default_content()

                                    nav.find_element(By.XPATH,'//*[@id="NETAModalDialog"]/div[1]/div[1]/button/span[1]').click()#DOCUMENTOS FILTRO
                                    time.sleep(1.5)

                                    nav.switch_to.frame('ifCruscottoUtenza')
                                    iframe_1 = WebDriverWait(nav, 10).until(
                                        EC.presence_of_element_located((By.ID,'ctl00_NetSiuCPH_lbl_crm_cruf_valstatofor'))#Status Fornecimento 
                                    )
                                    nav.find_element(By.ID,'ctl00_NetSiuCPH_rep_crm_desktop_utenza_ctl04_lb_crm_cruf_df_zoom').click()#PARCELAS
                                    time.sleep(1.5)
                                    nav.switch_to.default_content()

                                    nav.switch_to.frame('NETAModalDialogiFrame_1')
                                    iframe_1 = WebDriverWait(nav, 10).until(
                                        EC.presence_of_element_located((By.ID,'ctl00_NetSiuCPH_ctl01_ddl_crm_bparc_filtro'))#Status Fornecimento 
                                    )
                                    nav.find_element(By.ID,'ctl00_NetSiuCPH_ctl01_ddl_crm_bparc_filtro').click()#fILTOR
                                    time.sleep(1.5)

                                    nav.find_element(By.XPATH,'//*[@id="ctl00_NetSiuCPH_ctl01_ddl_crm_bparc_filtro"]/option[2]').click()#FECHAR
                                    time.sleep(3.0)


                                    try:
                                        tabela = WebDriverWait(nav, 10).until(
                                            EC.presence_of_element_located((By.ID, "ctl00_NetSiuCPH_ctl01_ngv_crm_bparc"))
                                        )
                                        n = 1
                                        while (n <= 1000):

                                            elemento = nav.find_element(By.ID,'ctl00_upProgress')

                                            if elemento.is_displayed():
                                                print("O elemento está visível na página.")

                                            else:
                                                print("O elemento não está visível na página.")
                                                break
                                                
                                            n += 1 

                                        palavra8 = "LOTE EFETIVADO"
                                        indice_coluna_desejada_Status = 18
                                        pelo_menos_uma_palavra_encontrada = False

                                        tabela_Status = nav.find_element(By.ID,'ctl00_NetSiuCPH_ctl01_ngv_crm_bparc')

                                        linhas_status = tabela_Status.find_elements(By.TAG_NAME,"tr")
                                        for linha in linhas_status:
                                            # Para cada linha, extraia as células (td)
                                            celulas = linha.find_elements(By.TAG_NAME, "td") 

                                            if len(celulas) > indice_coluna_desejada_Status:
                                                celula_desejada_Status = celulas[indice_coluna_desejada_Status]
                                                print(f"Conteúdo da coluna {indice_coluna_desejada_Status + 1}: {celula_desejada_Status.text}")

                                                if palavra8.lower() in celula_desejada_Status.text.lower() :
                                                    print(f"Pelo menos uma das palavras '{palavra8}' foi encontrada na Célula {indice_coluna_desejada_Status + 1} da linha.")
                                                    pelo_menos_uma_palavra_encontrada_Y = True
                                            
                                        if pelo_menos_uma_palavra_encontrada_Y:#FATURA ACORDO

                                            def encontrar_proxima_linha_vazia_na_coluna(sheet, coluna):
                                                linha = 1
                                                while sheet.cell(row=linha, column=coluna).value is not None:
                                                    linha += 1
                                                return linha

                                            try:
                                                wb = openpyxl.load_workbook('Consulta.xlsx')
                                            except FileNotFoundError:
                                                wb = openpyxl.Workbook()

                                            if 'Sheet' in wb.sheetnames:
                                                ws = wb['Sheet']
                                            else:
                                                ws = wb.active

                                            Hoje = datetime.now()
                                            Data = Hoje.strftime("%d/%m/%Y")
                                            Hora = Hoje.strftime("%H:%M:%S")    

                                            proxima_linha_b = encontrar_proxima_linha_vazia_na_coluna(ws, coluna=4)
                                            ws.cell(row=proxima_linha_b, column=4, value= texto )

                                            proxima_linha_C = encontrar_proxima_linha_vazia_na_coluna(ws, coluna=5)
                                            ws.cell(row=proxima_linha_C, column=5, value= "ACORDO EM ANDAMENTO")

                                            proxima_linha_D = encontrar_proxima_linha_vazia_na_coluna(ws, coluna=6)
                                            ws.cell(row=proxima_linha_D, column=6, value= Data )

                                            proxima_linha_E = encontrar_proxima_linha_vazia_na_coluna(ws, coluna=7)
                                            ws.cell(row=proxima_linha_E, column=7, value= Hora )
                                                            
                                            wb.save('Consulta.xlsx')

                                            nav.switch_to.default_content()

                                            nav.find_element(By.XPATH,'//*[@id="NETAModalDialog"]/div[1]/div[1]/button/span[1]').click()#DOCUMENTOS FILTRO
                                            time.sleep(1.5)
                                            
                                            nav.find_element(By.XPATH,'//*[@id="ctl00_NetSiuCPH_TabCRM_bli_tcrm_cruscotti"]/li/a/span/input').click()#FECHAR
                                            time.sleep(1.0)

                                            break

                                        else:

                                            def encontrar_proxima_linha_vazia_na_coluna(sheet, coluna):
                                                linha = 1
                                                while sheet.cell(row=linha, column=coluna).value is not None:
                                                    linha += 1
                                                return linha

                                            try:
                                                wb = openpyxl.load_workbook('Consulta.xlsx')
                                            except FileNotFoundError:
                                                wb = openpyxl.Workbook()

                                            if 'Sheet' in wb.sheetnames:
                                                ws = wb['Sheet']
                                            else:
                                                ws = wb.active

                                            Hoje = datetime.now()
                                            Data = Hoje.strftime("%d/%m/%Y")
                                            Hora = Hoje.strftime("%H:%M:%S")    

                                            proxima_linha_b = encontrar_proxima_linha_vazia_na_coluna(ws, coluna=4)
                                            ws.cell(row=proxima_linha_b, column=4, value= texto )

                                            proxima_linha_C = encontrar_proxima_linha_vazia_na_coluna(ws, coluna=5)
                                            ws.cell(row=proxima_linha_C, column=5, value= "DEBITO EM ABERTO - ACORDO ROMPIDO")

                                            proxima_linha_D = encontrar_proxima_linha_vazia_na_coluna(ws, coluna=6)
                                            ws.cell(row=proxima_linha_D, column=6, value= Data )

                                            proxima_linha_E = encontrar_proxima_linha_vazia_na_coluna(ws, coluna=7)
                                            ws.cell(row=proxima_linha_E, column=7, value= Hora )
                                                            
                                            wb.save('Consulta.xlsx')

                                            nav.switch_to.default_content()

                                            nav.find_element(By.XPATH,'//*[@id="NETAModalDialog"]/div[1]/div[1]/button/span[1]').click()#DOCUMENTOS FILTRO
                                            time.sleep(1.5)
                                            
                                            nav.find_element(By.XPATH,'//*[@id="ctl00_NetSiuCPH_TabCRM_bli_tcrm_cruscotti"]/li/a/span/input').click()#FECHAR
                                            time.sleep(1.0)

                                            break

                                    except TimeoutException:

                                        def encontrar_proxima_linha_vazia_na_coluna(sheet, coluna):
                                            linha = 1
                                            while sheet.cell(row=linha, column=coluna).value is not None:
                                                linha += 1
                                            return linha

                                        try:
                                            wb = openpyxl.load_workbook('Consulta.xlsx')
                                        except FileNotFoundError:
                                            wb = openpyxl.Workbook()

                                        if 'Sheet' in wb.sheetnames:
                                            ws = wb['Sheet']
                                        else:
                                            ws = wb.active

                                        Hoje = datetime.now()
                                        Data = Hoje.strftime("%d/%m/%Y")
                                        Hora = Hoje.strftime("%H:%M:%S")    

                                        proxima_linha_b = encontrar_proxima_linha_vazia_na_coluna(ws, coluna=4)
                                        ws.cell(row=proxima_linha_b, column=4, value= texto )

                                        proxima_linha_C = encontrar_proxima_linha_vazia_na_coluna(ws, coluna=5)
                                        ws.cell(row=proxima_linha_C, column=5, value= "CONTA PAGA")

                                        proxima_linha_D = encontrar_proxima_linha_vazia_na_coluna(ws, coluna=6)
                                        ws.cell(row=proxima_linha_D, column=6, value= Data )

                                        proxima_linha_E = encontrar_proxima_linha_vazia_na_coluna(ws, coluna=7)
                                        ws.cell(row=proxima_linha_E, column=7, value= Hora )
                                                        
                                        wb.save('Consulta.xlsx')

                                        nav.switch_to.default_content()

                                        nav.find_element(By.XPATH,'//*[@id="NETAModalDialog"]/div[1]/div[1]/button/span[1]').click()#DOCUMENTOS FILTRO
                                        time.sleep(1.5)
                                        
                                        nav.find_element(By.XPATH,'//*[@id="ctl00_NetSiuCPH_TabCRM_bli_tcrm_cruscotti"]/li/a/span/input').click()#FECHAR
                                        time.sleep(1.0)

                                        break

                                else:
                                    def encontrar_proxima_linha_vazia_na_coluna(sheet, coluna):
                                        linha = 1
                                        while sheet.cell(row=linha, column=coluna).value is not None:
                                            linha += 1
                                        return linha

                                    try:
                                        wb = openpyxl.load_workbook('Consulta.xlsx')
                                    except FileNotFoundError:
                                        wb = openpyxl.Workbook()

                                    if 'Sheet' in wb.sheetnames:
                                        ws = wb['Sheet']
                                    else:
                                        ws = wb.active

                                    Hoje = datetime.now()
                                    Data = Hoje.strftime("%d/%m/%Y")
                                    Hora = Hoje.strftime("%H:%M:%S")    

                                    proxima_linha_b = encontrar_proxima_linha_vazia_na_coluna(ws, coluna=4)
                                    ws.cell(row=proxima_linha_b, column=4, value= texto )

                                    proxima_linha_C = encontrar_proxima_linha_vazia_na_coluna(ws, coluna=5)
                                    ws.cell(row=proxima_linha_C, column=5, value= "CONTA PAGA")

                                    proxima_linha_D = encontrar_proxima_linha_vazia_na_coluna(ws, coluna=6)
                                    ws.cell(row=proxima_linha_D, column=6, value= Data )

                                    proxima_linha_E = encontrar_proxima_linha_vazia_na_coluna(ws, coluna=7)
                                    ws.cell(row=proxima_linha_E, column=7, value= Hora )
                                                    
                                    wb.save('Consulta.xlsx')

                                    nav.switch_to.default_content()

                                    nav.find_element(By.XPATH,'//*[@id="NETAModalDialog"]/div[1]/div[1]/button/span[1]').click()#DOCUMENTOS FILTRO
                                    time.sleep(1.5)
                                    
                                    nav.find_element(By.XPATH,'//*[@id="ctl00_NetSiuCPH_TabCRM_bli_tcrm_cruscotti"]/li/a/span/input').click()#FECHAR
                                    time.sleep(1.0)

                                    break
                        
                else:
                        print("palavra 1 nao encotrada-------------------------------------------------------------------------------")
                        print("ENTRANDO NA PALAVRA 2")
                        for linha in linhas:
                            # Para cada linha, extraia as células (td)
                            celulas = linha.find_elements(By.TAG_NAME, "td")  

                            if len(celulas) > indice_coluna_desejada:
                                celula_desejada = celulas[indice_coluna_desejada]
                                print(f"Conteúdo da coluna {indice_coluna_desejada + 1}: {celula_desejada.text}")

                                if palavra2.lower() in celula_desejada.text.lower() :
                                    print(f"Pelo menos uma das palavras '{palavra2}' foi encontrada na Célula {indice_coluna_desejada + 1} da linha.")

                                    indice_outra_coluna = 9

                                    if len(celulas) > indice_outra_coluna:
                                        outra_celula = celulas[indice_outra_coluna]
                                        OC_CONT += 1
                                        print(f"Conteúdo da outra coluna {indice_outra_coluna + 1}: {outra_celula.text}")

                                        if outra_celula.text == "N":
                                            print("Deu certo")

                                            OC += 1
                                            print(OC)

                                        else:
                                            OC_ERRO += 1
                                            print(OC_ERRO)
                    
                                    if OC > 0 :#OS
                                        def encontrar_proxima_linha_vazia_na_coluna(sheet, coluna):
                                            linha = 1
                                            while sheet.cell(row=linha, column=coluna).value is not None:
                                                linha += 1
                                            return linha

                                        try:
                                            wb = openpyxl.load_workbook('Consulta.xlsx')
                                        except FileNotFoundError:
                                            wb = openpyxl.Workbook()

                                        if 'Sheet' in wb.sheetnames:
                                            ws = wb['Sheet']
                                        else:
                                            ws = wb.active

                                        Hoje = datetime.now()
                                        Data = Hoje.strftime("%d/%m/%Y")
                                        Hora = Hoje.strftime("%H:%M:%S")  

                                        proxima_linha_b = encontrar_proxima_linha_vazia_na_coluna(ws, coluna=4)
                                        ws.cell(row=proxima_linha_b, column=4, value= texto )  

                                        proxima_linha_C = encontrar_proxima_linha_vazia_na_coluna(ws, coluna=5)
                                        ws.cell(row=proxima_linha_C, column=5, value= "DEBITO EM ABERTO - ORDEM DE CORTE")

                                        proxima_linha_D = encontrar_proxima_linha_vazia_na_coluna(ws, coluna=6)
                                        ws.cell(row=proxima_linha_D, column=6, value= Data )

                                        proxima_linha_E = encontrar_proxima_linha_vazia_na_coluna(ws, coluna=7)
                                        ws.cell(row=proxima_linha_E, column=7, value= Hora )
                                                            
                                        wb.save('Consulta.xlsx')

                                        nav.switch_to.default_content()
                                        nav.find_element(By.XPATH,'//*[@id="NETAModalDialog"]/div[1]/div[1]/button/span[1]').click()#DOCUMENTOS FILTRO
                                        time.sleep(1.5)
                                            
                                        nav.find_element(By.XPATH,'//*[@id="ctl00_NetSiuCPH_TabCRM_bli_tcrm_cruscotti"]/li/a/span/input').click()#FECHAR
                                        time.sleep(1.0)

                                        break
            
                                    elif OC_ERRO == OC_CONT:

                                            palavra7 = "FATURA ACORDO"
                                            indice_coluna_desejada_fatura = 2

                                            pelo_menos_uma_palavra_encontrada_X = False
                                            pelo_menos_uma_palavra_encontrada_Y = False

                                            linhas = tabela.find_elements(By.TAG_NAME,"tr")
                                            for linha in linhas:
                                                # Para cada linha, extraia as células (td)
                                                celulas = linha.find_elements(By.TAG_NAME, "td") 


                                                if len(celulas) > indice_coluna_desejada_fatura:
                                                    celula_desejada_fatura = celulas[indice_coluna_desejada_fatura]
                                                    print(f"Conteúdo da coluna {indice_coluna_desejada_fatura + 1}: {celula_desejada_fatura.text}")

                                                    if palavra7.lower() in celula_desejada_fatura.text.lower() :
                                                        print(f"Pelo menos uma das palavras '{palavra7}' foi encontrada na Célula {indice_coluna_desejada_fatura + 1} da linha.")
                                                        pelo_menos_uma_palavra_encontrada_X = True
                                                
                                            if pelo_menos_uma_palavra_encontrada_X:#FATURA ACORDO

                                                nav.switch_to.default_content()

                                                nav.find_element(By.XPATH,'//*[@id="NETAModalDialog"]/div[1]/div[1]/button/span[1]').click()#DOCUMENTOS FILTRO
                                                time.sleep(1.5)

                                                nav.switch_to.frame('ifCruscottoUtenza')
                                                iframe_1 = WebDriverWait(nav, 10).until(
                                                    EC.presence_of_element_located((By.ID,'ctl00_NetSiuCPH_lbl_crm_cruf_valstatofor'))#Status Fornecimento 
                                                )

                                                nav.find_element(By.ID,'ctl00_NetSiuCPH_rep_crm_desktop_utenza_ctl04_lb_crm_cruf_df_zoom').click()#PARCELAS
                                                time.sleep(1.5)
                                                nav.switch_to.default_content()

                                                nav.switch_to.frame('NETAModalDialogiFrame_1')
                                                iframe_1 = WebDriverWait(nav, 10).until(
                                                    EC.presence_of_element_located((By.ID,'ctl00_NetSiuCPH_ctl01_ddl_crm_bparc_filtro'))#Status Fornecimento 
                                                )
                                                nav.find_element(By.ID,'ctl00_NetSiuCPH_ctl01_ddl_crm_bparc_filtro').click()#fILTOR
                                                time.sleep(1.5)

                                                nav.find_element(By.XPATH,'//*[@id="ctl00_NetSiuCPH_ctl01_ddl_crm_bparc_filtro"]/option[2]').click()#FECHAR
                                                time.sleep(3.0)

                                                try:
                                                    tabela = WebDriverWait(nav, 10).until(
                                                        EC.presence_of_element_located((By.ID, "ctl00_NetSiuCPH_ctl01_ngv_crm_bparc"))
                                                    )
                                                    n = 1
                                                    while (n <= 1000):

                                                        elemento = nav.find_element(By.ID,'ctl00_upProgress')

                                                        if elemento.is_displayed():
                                                            print("O elemento está visível na página.")

                                                        else:
                                                            print("O elemento não está visível na página.")
                                                            break
                                                            
                                                        n += 1 

                                                    palavra8 = "LOTE EFETIVADO"
                                                    indice_coluna_desejada_Status = 18
                                                    pelo_menos_uma_palavra_encontrada = False

                                                    tabela_Status = nav.find_element(By.ID,'ctl00_NetSiuCPH_ctl01_ngv_crm_bparc')

                                                    linhas_status = tabela_Status.find_elements(By.TAG_NAME,"tr")
                                                    for linha in linhas_status:
                                                        # Para cada linha, extraia as células (td)
                                                        celulas = linha.find_elements(By.TAG_NAME, "td") 


                                                        if len(celulas) > indice_coluna_desejada_Status:
                                                            celula_desejada_Status = celulas[indice_coluna_desejada_Status]
                                                            print(f"Conteúdo da coluna {indice_coluna_desejada_Status + 1}: {celula_desejada_Status.text}")

                                                            if palavra8.lower() in celula_desejada_Status.text.lower() :
                                                                print(f"Pelo menos uma das palavras '{palavra8}' foi encontrada na Célula {indice_coluna_desejada_Status + 1} da linha.")
                                                                pelo_menos_uma_palavra_encontrada_Y = True
                                                        
                                                    if pelo_menos_uma_palavra_encontrada_Y:#FATURA ACORDO

                                                        def encontrar_proxima_linha_vazia_na_coluna(sheet, coluna):
                                                            linha = 1
                                                            while sheet.cell(row=linha, column=coluna).value is not None:
                                                                linha += 1
                                                            return linha

                                                        try:
                                                            wb = openpyxl.load_workbook('Consulta.xlsx')
                                                        except FileNotFoundError:
                                                            wb = openpyxl.Workbook()

                                                        if 'Sheet' in wb.sheetnames:
                                                            ws = wb['Sheet']
                                                        else:
                                                            ws = wb.active

                                                        Hoje = datetime.now()
                                                        Data = Hoje.strftime("%d/%m/%Y")
                                                        Hora = Hoje.strftime("%H:%M:%S")    

                                                        proxima_linha_b = encontrar_proxima_linha_vazia_na_coluna(ws, coluna=4)
                                                        ws.cell(row=proxima_linha_b, column=4, value= texto )

                                                        proxima_linha_C = encontrar_proxima_linha_vazia_na_coluna(ws, coluna=5)
                                                        ws.cell(row=proxima_linha_C, column=5, value= "ACORDO EM ANDAMENTO")

                                                        proxima_linha_D = encontrar_proxima_linha_vazia_na_coluna(ws, coluna=6)
                                                        ws.cell(row=proxima_linha_D, column=6, value= Data )

                                                        proxima_linha_E = encontrar_proxima_linha_vazia_na_coluna(ws, coluna=7)
                                                        ws.cell(row=proxima_linha_E, column=7, value= Hora )
                                                                        
                                                        wb.save('Consulta.xlsx')

                                                        nav.switch_to.default_content()

                                                        nav.find_element(By.XPATH,'//*[@id="NETAModalDialog"]/div[1]/div[1]/button/span[1]').click()#DOCUMENTOS FILTRO
                                                        time.sleep(1.5)
                                                        
                                                        nav.find_element(By.XPATH,'//*[@id="ctl00_NetSiuCPH_TabCRM_bli_tcrm_cruscotti"]/li/a/span/input').click()#FECHAR
                                                        time.sleep(1.0)

                                                        break

                                                        
                                                    else:

                                                        def encontrar_proxima_linha_vazia_na_coluna(sheet, coluna):
                                                            linha = 1
                                                            while sheet.cell(row=linha, column=coluna).value is not None:
                                                                linha += 1
                                                            return linha

                                                        try:
                                                            wb = openpyxl.load_workbook('Consulta.xlsx')
                                                        except FileNotFoundError:
                                                            wb = openpyxl.Workbook()

                                                        if 'Sheet' in wb.sheetnames:
                                                            ws = wb['Sheet']
                                                        else:
                                                            ws = wb.active

                                                        Hoje = datetime.now()
                                                        Data = Hoje.strftime("%d/%m/%Y")
                                                        Hora = Hoje.strftime("%H:%M:%S")    

                                                        proxima_linha_b = encontrar_proxima_linha_vazia_na_coluna(ws, coluna=4)
                                                        ws.cell(row=proxima_linha_b, column=4, value= texto )

                                                        proxima_linha_C = encontrar_proxima_linha_vazia_na_coluna(ws, coluna=5)
                                                        ws.cell(row=proxima_linha_C, column=5, value= "DEBITO EM ABERTO - ACORDO ROMPIDO")

                                                        proxima_linha_D = encontrar_proxima_linha_vazia_na_coluna(ws, coluna=6)
                                                        ws.cell(row=proxima_linha_D, column=6, value= Data )

                                                        proxima_linha_E = encontrar_proxima_linha_vazia_na_coluna(ws, coluna=7)
                                                        ws.cell(row=proxima_linha_E, column=7, value= Hora )
                                                                        
                                                        wb.save('Consulta.xlsx')

                                                        nav.switch_to.default_content()

                                                        nav.find_element(By.XPATH,'//*[@id="NETAModalDialog"]/div[1]/div[1]/button/span[1]').click()#DOCUMENTOS FILTRO
                                                        time.sleep(1.5)
                                                        
                                                        nav.find_element(By.XPATH,'//*[@id="ctl00_NetSiuCPH_TabCRM_bli_tcrm_cruscotti"]/li/a/span/input').click()#FECHAR
                                                        time.sleep(1.0)

                                                        break

                                                except TimeoutException:

                                                    def encontrar_proxima_linha_vazia_na_coluna(sheet, coluna):
                                                        linha = 1
                                                        while sheet.cell(row=linha, column=coluna).value is not None:
                                                            linha += 1
                                                        return linha

                                                    try:
                                                        wb = openpyxl.load_workbook('Consulta.xlsx')
                                                    except FileNotFoundError:
                                                        wb = openpyxl.Workbook()

                                                    if 'Sheet' in wb.sheetnames:
                                                        ws = wb['Sheet']
                                                    else:
                                                        ws = wb.active

                                                    Hoje = datetime.now()
                                                    Data = Hoje.strftime("%d/%m/%Y")
                                                    Hora = Hoje.strftime("%H:%M:%S")    

                                                    proxima_linha_b = encontrar_proxima_linha_vazia_na_coluna(ws, coluna=4)
                                                    ws.cell(row=proxima_linha_b, column=4, value= texto )

                                                    proxima_linha_C = encontrar_proxima_linha_vazia_na_coluna(ws, coluna=5)
                                                    ws.cell(row=proxima_linha_C, column=5, value= "CONTA PAGA")

                                                    proxima_linha_D = encontrar_proxima_linha_vazia_na_coluna(ws, coluna=6)
                                                    ws.cell(row=proxima_linha_D, column=6, value= Data )

                                                    proxima_linha_E = encontrar_proxima_linha_vazia_na_coluna(ws, coluna=7)
                                                    ws.cell(row=proxima_linha_E, column=7, value= Hora )
                                                                    
                                                    wb.save('Consulta.xlsx')

                                                    nav.switch_to.default_content()

                                                    nav.find_element(By.XPATH,'//*[@id="NETAModalDialog"]/div[1]/div[1]/button/span[1]').click()#DOCUMENTOS FILTRO
                                                    time.sleep(1.5)
                                                    
                                                    nav.find_element(By.XPATH,'//*[@id="ctl00_NetSiuCPH_TabCRM_bli_tcrm_cruscotti"]/li/a/span/input').click()#FECHAR
                                                    time.sleep(1.0)

                                                    break                                           
                                                
                                            else:
                                                def encontrar_proxima_linha_vazia_na_coluna(sheet, coluna):
                                                    linha = 1
                                                    while sheet.cell(row=linha, column=coluna).value is not None:
                                                        linha += 1
                                                    return linha

                                                try:
                                                    wb = openpyxl.load_workbook('Consulta.xlsx')
                                                except FileNotFoundError:
                                                    wb = openpyxl.Workbook()

                                                if 'Sheet' in wb.sheetnames:
                                                    ws = wb['Sheet']
                                                else:
                                                    ws = wb.active

                                                Hoje = datetime.now()
                                                Data = Hoje.strftime("%d/%m/%Y")
                                                Hora = Hoje.strftime("%H:%M:%S")    

                                                proxima_linha_b = encontrar_proxima_linha_vazia_na_coluna(ws, coluna=4)
                                                ws.cell(row=proxima_linha_b, column=4, value= texto )

                                                proxima_linha_C = encontrar_proxima_linha_vazia_na_coluna(ws, coluna=5)
                                                ws.cell(row=proxima_linha_C, column=5, value= "CONTA PAGA")

                                                proxima_linha_D = encontrar_proxima_linha_vazia_na_coluna(ws, coluna=6)
                                                ws.cell(row=proxima_linha_D, column=6, value= Data )

                                                proxima_linha_E = encontrar_proxima_linha_vazia_na_coluna(ws, coluna=7)
                                                ws.cell(row=proxima_linha_E, column=7, value= Hora )
                                                                
                                                wb.save('Consulta.xlsx')

                                                nav.switch_to.default_content()

                                                nav.find_element(By.XPATH,'//*[@id="NETAModalDialog"]/div[1]/div[1]/button/span[1]').click()#DOCUMENTOS FILTRO
                                                time.sleep(1.5)
                                                
                                                nav.find_element(By.XPATH,'//*[@id="ctl00_NetSiuCPH_TabCRM_bli_tcrm_cruscotti"]/li/a/span/input').click()#FECHAR
                                                time.sleep(1.0)

                                                break
                        
                        else:
                                print("palavra 2 nao encotrada-------------------------------------------------------------------------------")
                                for linha in linhas:
                                # Para cada linha, extraia as células (td)
                                    celulas = linha.find_elements(By.TAG_NAME, "td")  

                                    if len(celulas) > indice_coluna_desejada:
                                        celula_desejada = celulas[indice_coluna_desejada]
                                        print(f"Conteúdo da coluna {indice_coluna_desejada + 1}: {celula_desejada.text}")

                                        if palavra3.lower() in celula_desejada.text.lower() :
                                            print(f"Pelo menos uma das palavras '{palavra3}' foi encontrada na Célula {indice_coluna_desejada + 1} da linha.")
                                            #pelo_menos_uma_palavra_encontrada = True

                                            if  palavra6.lower() in celula_desejada.text.lower():

                                                def encontrar_proxima_linha_vazia_na_coluna(sheet, coluna):
                                                    linha = 1
                                                    while sheet.cell(row=linha, column=coluna).value is not None:
                                                        linha += 1
                                                    return linha

                                                try:
                                                    wb = openpyxl.load_workbook('Consulta.xlsx')
                                                except FileNotFoundError:
                                                    wb = openpyxl.Workbook()

                                                if 'Sheet' in wb.sheetnames:
                                                    ws = wb['Sheet']
                                                else:
                                                    ws = wb.active

                                                Hoje = datetime.now()
                                                Data = Hoje.strftime("%d/%m/%Y")
                                                Hora = Hoje.strftime("%H:%M:%S")    

                                                proxima_linha_b = encontrar_proxima_linha_vazia_na_coluna(ws, coluna=4)
                                                ws.cell(row=proxima_linha_b, column=4, value= texto )

                                                proxima_linha_C = encontrar_proxima_linha_vazia_na_coluna(ws, coluna=5)
                                                ws.cell(row=proxima_linha_C, column=5, value= "DEBITO EM ABERTO - JURÍDICO" )

                                                proxima_linha_D = encontrar_proxima_linha_vazia_na_coluna(ws, coluna=6)
                                                ws.cell(row=proxima_linha_D, column=6, value= Data )

                                                proxima_linha_E = encontrar_proxima_linha_vazia_na_coluna(ws, coluna=7)
                                                ws.cell(row=proxima_linha_E, column=7, value= Hora )
                                                                                            
                                                wb.save('Consulta.xlsx')

                                                nav.switch_to.default_content()
                                                nav.find_element(By.XPATH,'//*[@id="NETAModalDialog"]/div[1]/div[1]/button').click()#fECHARs
                                                time.sleep(1.5)
                                                nav.find_element(By.XPATH,'//*[@id="ctl00_NetSiuCPH_TabCRM_bli_tcrm_cruscotti"]/li/a/span/input').click()#FECHAR
                                                time.sleep(1.0)

                                                break

                                            else:                

                                                indice_outra_coluna = 9

                                                if len(celulas) > indice_outra_coluna:
                                                    outra_celula = celulas[indice_outra_coluna]
                                                    D_CONT += 1
                                                    print(f"Conteúdo da outra coluna {indice_outra_coluna + 1}: {outra_celula.text}")

                                                    if outra_celula.text == "N":
                                                        print("Deu certo")

                                                        D += 1
                                                        print(D)

                                                    else:
                                                        D_ERRO += 1
                                                        print(D_ERRO)
                                
                                                    if D > 0 :#OS
                                                        def encontrar_proxima_linha_vazia_na_coluna(sheet, coluna):
                                                            linha = 1
                                                            while sheet.cell(row=linha, column=coluna).value is not None:
                                                                linha += 1
                                                            return linha

                                                        try:
                                                            wb = openpyxl.load_workbook('Consulta.xlsx')
                                                        except FileNotFoundError:
                                                            wb = openpyxl.Workbook()

                                                        if 'Sheet' in wb.sheetnames:
                                                            ws = wb['Sheet']
                                                        else:
                                                            ws = wb.active

                                                        Hoje = datetime.now()
                                                        Data = Hoje.strftime("%d/%m/%Y")
                                                        Hora = Hoje.strftime("%H:%M:%S") 

                                                        proxima_linha_b = encontrar_proxima_linha_vazia_na_coluna(ws, coluna=4)
                                                        ws.cell(row=proxima_linha_b, column=4, value= texto )   

                                                        proxima_linha_C = encontrar_proxima_linha_vazia_na_coluna(ws, coluna=5)
                                                        ws.cell(row=proxima_linha_C, column=5, value= "DEBITO EM ABERTO - EXTRATO D")

                                                        proxima_linha_D = encontrar_proxima_linha_vazia_na_coluna(ws, coluna=6)
                                                        ws.cell(row=proxima_linha_D, column=6, value= Data )

                                                        proxima_linha_E = encontrar_proxima_linha_vazia_na_coluna(ws, coluna=7)
                                                        ws.cell(row=proxima_linha_E, column=7, value= Hora )
                                                                        
                                                        wb.save('Consulta.xlsx')

                                                        nav.switch_to.default_content()
                                                        nav.find_element(By.XPATH,'//*[@id="NETAModalDialog"]/div[1]/div[1]/button/span[1]').click()#DOCUMENTOS FILTRO
                                                        time.sleep(1.5)
                                                        
                                                        nav.find_element(By.XPATH,'//*[@id="ctl00_NetSiuCPH_TabCRM_bli_tcrm_cruscotti"]/li/a/span/input').click()#FECHAR
                                                        time.sleep(1.0)

                                                        break
                                                    
                                                    elif D_ERRO == D_CONT:

                                                        palavra7 = "FATURA ACORDO"
                                                        indice_coluna_desejada_fatura = 2

                                                        pelo_menos_uma_palavra_encontrada_X = False
                                                        pelo_menos_uma_palavra_encontrada_Y = False

                                                        linhas = tabela.find_elements(By.TAG_NAME,"tr")
                                                        for linha in linhas:
                                                            # Para cada linha, extraia as células (td)
                                                            celulas = linha.find_elements(By.TAG_NAME, "td") 


                                                            if len(celulas) > indice_coluna_desejada_fatura:
                                                                celula_desejada_fatura = celulas[indice_coluna_desejada_fatura]
                                                                print(f"Conteúdo da coluna {indice_coluna_desejada_fatura + 1}: {celula_desejada_fatura.text}")

                                                                if palavra7.lower() in celula_desejada_fatura.text.lower() :
                                                                    print(f"Pelo menos uma das palavras '{palavra7}' foi encontrada na Célula {indice_coluna_desejada_fatura + 1} da linha.")
                                                                    pelo_menos_uma_palavra_encontrada_X = True
                                                            
                                                        if pelo_menos_uma_palavra_encontrada_X:#FATURA ACORDO

                                                            nav.switch_to.default_content()

                                                            nav.find_element(By.XPATH,'//*[@id="NETAModalDialog"]/div[1]/div[1]/button/span[1]').click()#DOCUMENTOS FILTRO
                                                            time.sleep(1.5)

                                                            nav.switch_to.frame('ifCruscottoUtenza')
                                                            iframe_1 = WebDriverWait(nav, 10).until(
                                                                EC.presence_of_element_located((By.ID,'ctl00_NetSiuCPH_lbl_crm_cruf_valstatofor'))#Status Fornecimento 
                                                            )
                                                            nav.find_element(By.ID,'ctl00_NetSiuCPH_rep_crm_desktop_utenza_ctl04_lb_crm_cruf_df_zoom').click()#PARCELAS
                                                            time.sleep(1.5)
                                                            nav.switch_to.default_content()

                                                            nav.switch_to.frame('NETAModalDialogiFrame_1')
                                                            iframe_1 = WebDriverWait(nav, 10).until(
                                                                EC.presence_of_element_located((By.ID,'ctl00_NetSiuCPH_ctl01_ddl_crm_bparc_filtro'))#Status Fornecimento 
                                                            )

                                                            nav.find_element(By.ID,'ctl00_NetSiuCPH_ctl01_ddl_crm_bparc_filtro').click()#fILTOR
                                                            time.sleep(1.5)

                                                            nav.find_element(By.XPATH,'//*[@id="ctl00_NetSiuCPH_ctl01_ddl_crm_bparc_filtro"]/option[2]').click()#FECHAR
                                                            time.sleep(3.0)

                                                            try:
                                                                tabela = WebDriverWait(nav, 10).until(
                                                                    EC.presence_of_element_located((By.ID, "ctl00_NetSiuCPH_ctl01_ngv_crm_bparc"))
                                                                )
                                                                n = 1
                                                                while (n <= 1000):

                                                                    elemento = nav.find_element(By.ID,'ctl00_upProgress')

                                                                    if elemento.is_displayed():
                                                                        print("O elemento está visível na página.")

                                                                    else:
                                                                        print("O elemento não está visível na página.")
                                                                        break
                                                                        
                                                                    n += 1 

                                                                palavra8 = "LOTE EFETIVADO"
                                                                indice_coluna_desejada_Status = 18
                                                                pelo_menos_uma_palavra_encontrada = False

                                                                tabela_Status = nav.find_element(By.ID,'ctl00_NetSiuCPH_ctl01_ngv_crm_bparc')

                                                                linhas_status = tabela_Status.find_elements(By.TAG_NAME,"tr")
                                                                for linha in linhas_status:
                                                                    # Para cada linha, extraia as células (td)
                                                                    celulas = linha.find_elements(By.TAG_NAME, "td") 


                                                                    if len(celulas) > indice_coluna_desejada_Status:
                                                                        celula_desejada_Status = celulas[indice_coluna_desejada_Status]
                                                                        print(f"Conteúdo da coluna {indice_coluna_desejada_Status + 1}: {celula_desejada_Status.text}")

                                                                        if palavra8.lower() in celula_desejada_Status.text.lower() :
                                                                            print(f"Pelo menos uma das palavras '{palavra8}' foi encontrada na Célula {indice_coluna_desejada_Status + 1} da linha.")
                                                                            pelo_menos_uma_palavra_encontrada_Y = True
                                                                    
                                                                if pelo_menos_uma_palavra_encontrada_Y:#FATURA ACORDO

                                                                    def encontrar_proxima_linha_vazia_na_coluna(sheet, coluna):
                                                                        linha = 1
                                                                        while sheet.cell(row=linha, column=coluna).value is not None:
                                                                            linha += 1
                                                                        return linha

                                                                    try:
                                                                        wb = openpyxl.load_workbook('Consulta.xlsx')
                                                                    except FileNotFoundError:
                                                                        wb = openpyxl.Workbook()

                                                                    if 'Sheet' in wb.sheetnames:
                                                                        ws = wb['Sheet']
                                                                    else:
                                                                        ws = wb.active

                                                                    Hoje = datetime.now()
                                                                    Data = Hoje.strftime("%d/%m/%Y")
                                                                    Hora = Hoje.strftime("%H:%M:%S")    


                                                                    proxima_linha_b = encontrar_proxima_linha_vazia_na_coluna(ws, coluna=4)
                                                                    ws.cell(row=proxima_linha_b, column=4, value= texto )

                                                                    proxima_linha_C = encontrar_proxima_linha_vazia_na_coluna(ws, coluna=5)
                                                                    ws.cell(row=proxima_linha_C, column=5, value= "ACORDO EM ANDAMENTO")

                                                                    proxima_linha_D = encontrar_proxima_linha_vazia_na_coluna(ws, coluna=6)
                                                                    ws.cell(row=proxima_linha_D, column=6, value= Data )

                                                                    proxima_linha_E = encontrar_proxima_linha_vazia_na_coluna(ws, coluna=7)
                                                                    ws.cell(row=proxima_linha_E, column=7, value= Hora )
                                                                                    
                                                                    wb.save('Consulta.xlsx')

                                                                    nav.switch_to.default_content()

                                                                    nav.find_element(By.XPATH,'//*[@id="NETAModalDialog"]/div[1]/div[1]/button/span[1]').click()#DOCUMENTOS FILTRO
                                                                    time.sleep(1.5)
                                                                    
                                                                    nav.find_element(By.XPATH,'//*[@id="ctl00_NetSiuCPH_TabCRM_bli_tcrm_cruscotti"]/li/a/span/input').click()#FECHAR
                                                                    time.sleep(1.0)

                                                                    break
                                                                    
                                                                else:

                                                                    def encontrar_proxima_linha_vazia_na_coluna(sheet, coluna):
                                                                        linha = 1
                                                                        while sheet.cell(row=linha, column=coluna).value is not None:
                                                                            linha += 1
                                                                        return linha

                                                                    try:
                                                                        wb = openpyxl.load_workbook('Consulta.xlsx')
                                                                    except FileNotFoundError:
                                                                        wb = openpyxl.Workbook()

                                                                    if 'Sheet' in wb.sheetnames:
                                                                        ws = wb['Sheet']
                                                                    else:
                                                                        ws = wb.active

                                                                    Hoje = datetime.now()
                                                                    Data = Hoje.strftime("%d/%m/%Y")
                                                                    Hora = Hoje.strftime("%H:%M:%S")   

                                                                    proxima_linha_b = encontrar_proxima_linha_vazia_na_coluna(ws, coluna=4)
                                                                    ws.cell(row=proxima_linha_b, column=4, value= texto ) 

                                                                    proxima_linha_C = encontrar_proxima_linha_vazia_na_coluna(ws, coluna=5)
                                                                    ws.cell(row=proxima_linha_C, column=5, value= "DEBITO EM ABERTO - ACORDO ROMPIDO")

                                                                    proxima_linha_D = encontrar_proxima_linha_vazia_na_coluna(ws, coluna=6)
                                                                    ws.cell(row=proxima_linha_D, column=6, value= Data )

                                                                    proxima_linha_E = encontrar_proxima_linha_vazia_na_coluna(ws, coluna=7)
                                                                    ws.cell(row=proxima_linha_E, column=7, value= Hora )
                                                                                    
                                                                    wb.save('Consulta.xlsx')

                                                                    nav.switch_to.default_content()

                                                                    nav.find_element(By.XPATH,'//*[@id="NETAModalDialog"]/div[1]/div[1]/button/span[1]').click()#DOCUMENTOS FILTRO
                                                                    time.sleep(1.5)
                                                                    
                                                                    nav.find_element(By.XPATH,'//*[@id="ctl00_NetSiuCPH_TabCRM_bli_tcrm_cruscotti"]/li/a/span/input').click()#FECHAR
                                                                    time.sleep(1.0)

                                                                    break
                                                            
                                                            except TimeoutException:

                                                                def encontrar_proxima_linha_vazia_na_coluna(sheet, coluna):
                                                                    linha = 1
                                                                    while sheet.cell(row=linha, column=coluna).value is not None:
                                                                        linha += 1
                                                                    return linha

                                                                try:
                                                                    wb = openpyxl.load_workbook('Consulta.xlsx')
                                                                except FileNotFoundError:
                                                                    wb = openpyxl.Workbook()

                                                                if 'Sheet' in wb.sheetnames:
                                                                    ws = wb['Sheet']
                                                                else:
                                                                    ws = wb.active

                                                                Hoje = datetime.now()
                                                                Data = Hoje.strftime("%d/%m/%Y")
                                                                Hora = Hoje.strftime("%H:%M:%S")    

                                                                proxima_linha_b = encontrar_proxima_linha_vazia_na_coluna(ws, coluna=4)
                                                                ws.cell(row=proxima_linha_b, column=4, value= texto )

                                                                proxima_linha_C = encontrar_proxima_linha_vazia_na_coluna(ws, coluna=5)
                                                                ws.cell(row=proxima_linha_C, column=5, value= "CONTA PAGA")

                                                                proxima_linha_D = encontrar_proxima_linha_vazia_na_coluna(ws, coluna=6)
                                                                ws.cell(row=proxima_linha_D, column=6, value= Data )

                                                                proxima_linha_E = encontrar_proxima_linha_vazia_na_coluna(ws, coluna=7)
                                                                ws.cell(row=proxima_linha_E, column=7, value= Hora )
                                                                                
                                                                wb.save('Consulta.xlsx')

                                                                nav.switch_to.default_content()

                                                                nav.find_element(By.XPATH,'//*[@id="NETAModalDialog"]/div[1]/div[1]/button/span[1]').click()#DOCUMENTOS FILTRO
                                                                time.sleep(1.5)
                                                                
                                                                nav.find_element(By.XPATH,'//*[@id="ctl00_NetSiuCPH_TabCRM_bli_tcrm_cruscotti"]/li/a/span/input').click()#FECHAR
                                                                time.sleep(1.0)

                                                                break        

                                                    else:
                                                            def encontrar_proxima_linha_vazia_na_coluna(sheet, coluna):
                                                                linha = 1
                                                                while sheet.cell(row=linha, column=coluna).value is not None:
                                                                    linha += 1
                                                                return linha

                                                            try:
                                                                wb = openpyxl.load_workbook('Consulta.xlsx')
                                                            except FileNotFoundError:
                                                                wb = openpyxl.Workbook()

                                                            if 'Sheet' in wb.sheetnames:
                                                                ws = wb['Sheet']
                                                            else:
                                                                ws = wb.active

                                                            Hoje = datetime.now()
                                                            Data = Hoje.strftime("%d/%m/%Y")
                                                            Hora = Hoje.strftime("%H:%M:%S")    

                                                            proxima_linha_b = encontrar_proxima_linha_vazia_na_coluna(ws, coluna=4)
                                                            ws.cell(row=proxima_linha_b, column=4, value= texto )

                                                            proxima_linha_C = encontrar_proxima_linha_vazia_na_coluna(ws, coluna=5)
                                                            ws.cell(row=proxima_linha_C, column=5, value= "CONTA PAGA")

                                                            proxima_linha_D = encontrar_proxima_linha_vazia_na_coluna(ws, coluna=6)
                                                            ws.cell(row=proxima_linha_D, column=6, value= Data )

                                                            proxima_linha_E = encontrar_proxima_linha_vazia_na_coluna(ws, coluna=7)
                                                            ws.cell(row=proxima_linha_E, column=7, value= Hora )
                                                                            
                                                            wb.save('Consulta.xlsx')

                                                            nav.switch_to.default_content()

                                                            nav.find_element(By.XPATH,'//*[@id="NETAModalDialog"]/div[1]/div[1]/button/span[1]').click()#DOCUMENTOS FILTRO
                                                            time.sleep(1.5)
                                                            
                                                            nav.find_element(By.XPATH,'//*[@id="ctl00_NetSiuCPH_TabCRM_bli_tcrm_cruscotti"]/li/a/span/input').click()#FECHAR
                                                            time.sleep(1.0)

                                                            break
                        
                                else:
                                        print("palavra 3 nao encotrada-------------------------------------------------------------------------------")

                                        linhas = tabela.find_elements(By.TAG_NAME,"tr")
                                        for linha in linhas:
                                            # Para cada linha, extraia as células (td)
                                            celulas = linha.find_elements(By.TAG_NAME, "td")  
                                            
                                            if len(celulas) > indice_coluna_desejada:
                                                celula_desejada = celulas[indice_coluna_desejada]
                                                
                                                print(f"Conteúdo da coluna {indice_coluna_desejada + 1}: {celula_desejada.text} ")

                                                if palavra4.lower() in celula_desejada.text.lower():
                                                    print(f"Pelo menos uma das palavras '{palavra4}' foi encontrada na Célula {indice_coluna_desejada + 1} da linha.")
                                                    #pelo_menos_uma_palavra_encontrada = True 

                                                    if palavra5.lower() in celula_desejada.text.lower():

                                                        def encontrar_proxima_linha_vazia_na_coluna(sheet, coluna):
                                                            linha = 1
                                                            while sheet.cell(row=linha, column=coluna).value is not None:
                                                                linha += 1
                                                            return linha

                                                        try:
                                                            wb = openpyxl.load_workbook('Consulta.xlsx')
                                                        except FileNotFoundError:
                                                            wb = openpyxl.Workbook()

                                                        if 'Sheet' in wb.sheetnames:
                                                            ws = wb['Sheet']
                                                        else:
                                                            ws = wb.active

                                                        Hoje = datetime.now()
                                                        Data = Hoje.strftime("%d/%m/%Y")
                                                        Hora = Hoje.strftime("%H:%M:%S")    

                                                        proxima_linha_b = encontrar_proxima_linha_vazia_na_coluna(ws, coluna=4)
                                                        ws.cell(row=proxima_linha_b, column=4, value= texto )

                                                        proxima_linha_C = encontrar_proxima_linha_vazia_na_coluna(ws, coluna=5)
                                                        ws.cell(row=proxima_linha_C, column=5, value= "DEBITO EM ABERTO - JURÍDICO" )

                                                        proxima_linha_D = encontrar_proxima_linha_vazia_na_coluna(ws, coluna=6)
                                                        ws.cell(row=proxima_linha_D, column=6, value= Data )

                                                        proxima_linha_E = encontrar_proxima_linha_vazia_na_coluna(ws, coluna=7)
                                                        ws.cell(row=proxima_linha_E, column=7, value= Hora )
                                                                                            
                                                        wb.save('Consulta.xlsx')

                                                        nav.switch_to.default_content()
                                                        nav.find_element(By.XPATH,'//*[@id="NETAModalDialog"]/div[1]/div[1]/button').click()#fECHARs
                                                        time.sleep(1.5)
                                                        nav.find_element(By.XPATH,'//*[@id="ctl00_NetSiuCPH_TabCRM_bli_tcrm_cruscotti"]/li/a/span/input').click()#FECHAR
                                                        time.sleep(1.0)

                                                        break

                                                    else:

                                                        indice_outra_coluna = 9

                                                        if len(celulas) > indice_outra_coluna:
                                                            outra_celula = celulas[indice_outra_coluna]
                                                            C_CONT += 1
                                                            print(f"Conteúdo da outra coluna {indice_outra_coluna + 1}: {outra_celula.text}: {OS_CONT}")

                                                            if outra_celula.text == "N":
                                                                print("Deu certo")

                                                                C += 1
                                                                print(C)
                                                            else:
                                                                C_ERRO += 1
                                                                print(C_ERRO)            
                                                                        
                                                        if C > 0 :#OS

                                                            def encontrar_proxima_linha_vazia_na_coluna(sheet, coluna):
                                                                linha = 1
                                                                while sheet.cell(row=linha, column=coluna).value is not None:
                                                                    linha += 1
                                                                return linha

                                                            try:
                                                                wb = openpyxl.load_workbook('Consulta.xlsx')
                                                            except FileNotFoundError:
                                                                wb = openpyxl.Workbook()

                                                            if 'Sheet' in wb.sheetnames:
                                                                ws = wb['Sheet']
                                                            else:
                                                                ws = wb.active

                                                            Hoje = datetime.now()
                                                            Data = Hoje.strftime("%d/%m/%Y")
                                                            Hora = Hoje.strftime("%H:%M:%S")    

                                                            proxima_linha_b = encontrar_proxima_linha_vazia_na_coluna(ws, coluna=4)
                                                            ws.cell(row=proxima_linha_b, column=4, value= texto )

                                                            proxima_linha_C = encontrar_proxima_linha_vazia_na_coluna(ws, coluna=5)
                                                            ws.cell(row=proxima_linha_C, column=5, value= "DEBITO EM ABERTO - EXTRATO C" )

                                                            proxima_linha_D = encontrar_proxima_linha_vazia_na_coluna(ws, coluna=6)
                                                            ws.cell(row=proxima_linha_D, column=6, value= Data )

                                                            proxima_linha_E = encontrar_proxima_linha_vazia_na_coluna(ws, coluna=7)
                                                            ws.cell(row=proxima_linha_E, column=7, value= Hora )
                                                                            
                                                            wb.save('Consulta.xlsx')

                                                            nav.switch_to.default_content()
                                                            nav.find_element(By.XPATH,'//*[@id="NETAModalDialog"]/div[1]/div[1]/button').click()#fECHARs
                                                            time.sleep(1.5)
                                                            nav.find_element(By.XPATH,'//*[@id="ctl00_NetSiuCPH_TabCRM_bli_tcrm_cruscotti"]/li/a/span/input').click()#FECHAR
                                                            time.sleep(1.0)

                                                            break

                                                        elif C_ERRO == C_CONT:

                                                            palavra7 = "FATURA ACORDO"
                                                            indice_coluna_desejada_fatura = 2

                                                            pelo_menos_uma_palavra_encontrada_X = False
                                                            pelo_menos_uma_palavra_encontrada_Y = False

                                                            linhas = tabela.find_elements(By.TAG_NAME,"tr")
                                                            for linha in linhas:
                                                                # Para cada linha, extraia as células (td)
                                                                celulas = linha.find_elements(By.TAG_NAME, "td") 


                                                                if len(celulas) > indice_coluna_desejada_fatura:
                                                                    celula_desejada_fatura = celulas[indice_coluna_desejada_fatura]
                                                                    print(f"Conteúdo da coluna {indice_coluna_desejada_fatura + 1}: {celula_desejada_fatura.text}")

                                                                    if palavra7.lower() in celula_desejada_fatura.text.lower() :
                                                                        print(f"Pelo menos uma das palavras '{palavra7}' foi encontrada na Célula {indice_coluna_desejada_fatura + 1} da linha.")
                                                                        pelo_menos_uma_palavra_encontrada_X = True
                                                                
                                                            if pelo_menos_uma_palavra_encontrada_X:#FATURA ACORDO

                                                                nav.switch_to.default_content()

                                                                nav.find_element(By.XPATH,'//*[@id="NETAModalDialog"]/div[1]/div[1]/button/span[1]').click()#DOCUMENTOS FILTRO
                                                                time.sleep(1.5)

                                                                nav.switch_to.frame('ifCruscottoUtenza')
                                                                iframe_1 = WebDriverWait(nav, 10).until(
                                                                    EC.presence_of_element_located((By.ID,'ctl00_NetSiuCPH_lbl_crm_cruf_valstatofor'))#Status Fornecimento 
                                                                )
                                                                nav.find_element(By.ID,'ctl00_NetSiuCPH_rep_crm_desktop_utenza_ctl04_lb_crm_cruf_df_zoom').click()#PARCELAS
                                                                time.sleep(1.5)
                                                                nav.switch_to.default_content()

                                                                nav.switch_to.frame('NETAModalDialogiFrame_1')
                                                                iframe_1 = WebDriverWait(nav, 10).until(
                                                                    EC.presence_of_element_located((By.ID,'ctl00_NetSiuCPH_ctl01_ddl_crm_bparc_filtro'))#Status Fornecimento 
                                                                )
                                                                nav.find_element(By.ID,'ctl00_NetSiuCPH_ctl01_ddl_crm_bparc_filtro').click()#fILTOR
                                                                time.sleep(1.5)

                                                                nav.find_element(By.XPATH,'//*[@id="ctl00_NetSiuCPH_ctl01_ddl_crm_bparc_filtro"]/option[2]').click()#FECHAR
                                                                time.sleep(3.0)

                                                                try:
                                                                    tabela = WebDriverWait(nav, 10).until(
                                                                        EC.presence_of_element_located((By.ID, "ctl00_NetSiuCPH_ctl01_ngv_crm_bparc"))
                                                                    )
                                                                    n = 1
                                                                    while (n <= 1000):

                                                                        elemento = nav.find_element(By.ID,'ctl00_upProgress')

                                                                        if elemento.is_displayed():
                                                                            print("O elemento está visível na página.")

                                                                        else:
                                                                            print("O elemento não está visível na página.")
                                                                            break
                                                                            
                                                                        n += 1 

                                                                    palavra8 = "LOTE EFETIVADO"
                                                                    indice_coluna_desejada_Status = 18
                                                                    pelo_menos_uma_palavra_encontrada = False

                                                                    tabela_Status = nav.find_element(By.ID,'ctl00_NetSiuCPH_ctl01_ngv_crm_bparc')

                                                                    linhas_status = tabela_Status.find_elements(By.TAG_NAME,"tr")
                                                                    for linha in linhas_status:
                                                                        # Para cada linha, extraia as células (td)
                                                                        celulas = linha.find_elements(By.TAG_NAME, "td") 

                                                                        if len(celulas) > indice_coluna_desejada_Status:
                                                                            celula_desejada_Status = celulas[indice_coluna_desejada_Status]
                                                                            print(f"Conteúdo da coluna {indice_coluna_desejada_Status + 1}: {celula_desejada_Status.text}")

                                                                            if palavra8.lower() in celula_desejada_Status.text.lower() :
                                                                                print(f"Pelo menos uma das palavras '{palavra8}' foi encontrada na Célula {indice_coluna_desejada_Status + 1} da linha.")
                                                                                pelo_menos_uma_palavra_encontrada_Y = True
                                                                        
                                                                    if pelo_menos_uma_palavra_encontrada_Y:#FATURA ACORDO

                                                                        def encontrar_proxima_linha_vazia_na_coluna(sheet, coluna):
                                                                            linha = 1
                                                                            while sheet.cell(row=linha, column=coluna).value is not None:
                                                                                linha += 1
                                                                            return linha

                                                                        try:
                                                                            wb = openpyxl.load_workbook('Consulta.xlsx')
                                                                        except FileNotFoundError:
                                                                            wb = openpyxl.Workbook()

                                                                        if 'Sheet' in wb.sheetnames:
                                                                            ws = wb['Sheet']
                                                                        else:
                                                                            ws = wb.active

                                                                        Hoje = datetime.now()
                                                                        Data = Hoje.strftime("%d/%m/%Y")
                                                                        Hora = Hoje.strftime("%H:%M:%S")    

                                                                        proxima_linha_b = encontrar_proxima_linha_vazia_na_coluna(ws, coluna=4)
                                                                        ws.cell(row=proxima_linha_b, column=4, value= texto )

                                                                        proxima_linha_C = encontrar_proxima_linha_vazia_na_coluna(ws, coluna=5)
                                                                        ws.cell(row=proxima_linha_C, column=5, value= "ACORDO EM ANDAMENTO")

                                                                        proxima_linha_D = encontrar_proxima_linha_vazia_na_coluna(ws, coluna=6)
                                                                        ws.cell(row=proxima_linha_D, column=6, value= Data )

                                                                        proxima_linha_E = encontrar_proxima_linha_vazia_na_coluna(ws, coluna=7)
                                                                        ws.cell(row=proxima_linha_E, column=7, value= Hora )
                                                                                        
                                                                        wb.save('Consulta.xlsx')

                                                                        nav.switch_to.default_content()

                                                                        nav.find_element(By.XPATH,'//*[@id="NETAModalDialog"]/div[1]/div[1]/button/span[1]').click()#DOCUMENTOS FILTRO
                                                                        time.sleep(1.5)
                                                                        
                                                                        nav.find_element(By.XPATH,'//*[@id="ctl00_NetSiuCPH_TabCRM_bli_tcrm_cruscotti"]/li/a/span/input').click()#FECHAR
                                                                        time.sleep(1.0)

                                                                        break
                                                                    
                                                                    else:

                                                                        def encontrar_proxima_linha_vazia_na_coluna(sheet, coluna):
                                                                            linha = 1
                                                                            while sheet.cell(row=linha, column=coluna).value is not None:
                                                                                linha += 1
                                                                            return linha

                                                                        try:
                                                                            wb = openpyxl.load_workbook('Consulta.xlsx')
                                                                        except FileNotFoundError:
                                                                            wb = openpyxl.Workbook()

                                                                        if 'Sheet' in wb.sheetnames:
                                                                            ws = wb['Sheet']
                                                                        else:
                                                                            ws = wb.active

                                                                        Hoje = datetime.now()
                                                                        Data = Hoje.strftime("%d/%m/%Y")
                                                                        Hora = Hoje.strftime("%H:%M:%S") 

                                                                        proxima_linha_b = encontrar_proxima_linha_vazia_na_coluna(ws, coluna=4)
                                                                        ws.cell(row=proxima_linha_b, column=4, value= texto )   

                                                                        proxima_linha_C = encontrar_proxima_linha_vazia_na_coluna(ws, coluna=5)
                                                                        ws.cell(row=proxima_linha_C, column=5, value= "DEBITO EM ABERTO - ACORDO ROMPIDO")

                                                                        proxima_linha_D = encontrar_proxima_linha_vazia_na_coluna(ws, coluna=6)
                                                                        ws.cell(row=proxima_linha_D, column=6, value= Data )

                                                                        proxima_linha_E = encontrar_proxima_linha_vazia_na_coluna(ws, coluna=7)
                                                                        ws.cell(row=proxima_linha_E, column=7, value= Hora )
                                                                                        
                                                                        wb.save('Consulta.xlsx')

                                                                        nav.switch_to.default_content()

                                                                        nav.find_element(By.XPATH,'//*[@id="NETAModalDialog"]/div[1]/div[1]/button/span[1]').click()#DOCUMENTOS FILTRO
                                                                        time.sleep(1.5)
                                                                        
                                                                        nav.find_element(By.XPATH,'//*[@id="ctl00_NetSiuCPH_TabCRM_bli_tcrm_cruscotti"]/li/a/span/input').click()#FECHAR
                                                                        time.sleep(1.0)

                                                                        break
                                                                
                                                                except TimeoutException:

                                                                    def encontrar_proxima_linha_vazia_na_coluna(sheet, coluna):
                                                                        linha = 1
                                                                        while sheet.cell(row=linha, column=coluna).value is not None:
                                                                            linha += 1
                                                                        return linha

                                                                    try:
                                                                        wb = openpyxl.load_workbook('Consulta.xlsx')
                                                                    except FileNotFoundError:
                                                                        wb = openpyxl.Workbook()

                                                                    if 'Sheet' in wb.sheetnames:
                                                                        ws = wb['Sheet']
                                                                    else:
                                                                        ws = wb.active

                                                                    Hoje = datetime.now()
                                                                    Data = Hoje.strftime("%d/%m/%Y")
                                                                    Hora = Hoje.strftime("%H:%M:%S")    

                                                                    proxima_linha_b = encontrar_proxima_linha_vazia_na_coluna(ws, coluna=4)
                                                                    ws.cell(row=proxima_linha_b, column=4, value= texto )

                                                                    proxima_linha_C = encontrar_proxima_linha_vazia_na_coluna(ws, coluna=5)
                                                                    ws.cell(row=proxima_linha_C, column=5, value= "CONTA PAGA")

                                                                    proxima_linha_D = encontrar_proxima_linha_vazia_na_coluna(ws, coluna=6)
                                                                    ws.cell(row=proxima_linha_D, column=6, value= Data )

                                                                    proxima_linha_E = encontrar_proxima_linha_vazia_na_coluna(ws, coluna=7)
                                                                    ws.cell(row=proxima_linha_E, column=7, value= Hora )
                                                                                    
                                                                    wb.save('Consulta.xlsx')

                                                                    nav.switch_to.default_content()

                                                                    nav.find_element(By.XPATH,'//*[@id="NETAModalDialog"]/div[1]/div[1]/button/span[1]').click()#DOCUMENTOS FILTRO
                                                                    time.sleep(1.5)
                                                                    
                                                                    nav.find_element(By.XPATH,'//*[@id="ctl00_NetSiuCPH_TabCRM_bli_tcrm_cruscotti"]/li/a/span/input').click()#FECHAR
                                                                    time.sleep(1.0)

                                                                    break

                                                            else:
                                                                def encontrar_proxima_linha_vazia_na_coluna(sheet, coluna):
                                                                    linha = 1
                                                                    while sheet.cell(row=linha, column=coluna).value is not None:
                                                                        linha += 1
                                                                    return linha

                                                                try:
                                                                    wb = openpyxl.load_workbook('Consulta.xlsx')
                                                                except FileNotFoundError:
                                                                    wb = openpyxl.Workbook()

                                                                if 'Sheet' in wb.sheetnames:
                                                                    ws = wb['Sheet']
                                                                else:
                                                                    ws = wb.active

                                                                Hoje = datetime.now()
                                                                Data = Hoje.strftime("%d/%m/%Y")
                                                                Hora = Hoje.strftime("%H:%M:%S")    

                                                                proxima_linha_b = encontrar_proxima_linha_vazia_na_coluna(ws, coluna=4)
                                                                ws.cell(row=proxima_linha_b, column=4, value= texto )

                                                                proxima_linha_C = encontrar_proxima_linha_vazia_na_coluna(ws, coluna=5)
                                                                ws.cell(row=proxima_linha_C, column=5, value= "CONTA PAGA")

                                                                proxima_linha_D = encontrar_proxima_linha_vazia_na_coluna(ws, coluna=6)
                                                                ws.cell(row=proxima_linha_D, column=6, value= Data )

                                                                proxima_linha_E = encontrar_proxima_linha_vazia_na_coluna(ws, coluna=7)
                                                                ws.cell(row=proxima_linha_E, column=7, value= Hora )
                                                                                
                                                                wb.save('Consulta.xlsx')

                                                                nav.switch_to.default_content()

                                                                nav.find_element(By.XPATH,'//*[@id="NETAModalDialog"]/div[1]/div[1]/button/span[1]').click()#DOCUMENTOS FILTRO
                                                                time.sleep(1.5)
                                                                
                                                                nav.find_element(By.XPATH,'//*[@id="ctl00_NetSiuCPH_TabCRM_bli_tcrm_cruscotti"]/li/a/span/input').click()#FECHAR
                                                                time.sleep(1.0)

                                                                break

                                        else:
                                            n = 1
                                            while (n <= 1000):

                                                elemento = nav.find_element(By.ID,'ctl00_upProgress')

                                                if elemento.is_displayed():
                                                    print("O elemento está visível na página.")

                                                else:
                                                    print("O elemento não está visível na página.")
                                                    break
                                                            
                                                n += 1 

                                            palavra7 = "FATURA ACORDO"
                                            indice_coluna_desejada_fatura = 2

                                            pelo_menos_uma_palavra_encontrada_X = False
                                            pelo_menos_uma_palavra_encontrada_Y = False

                                            linhas = tabela.find_elements(By.TAG_NAME,"tr")
                                            for linha in linhas:
                                                # Para cada linha, extraia as células (td)
                                                celulas = linha.find_elements(By.TAG_NAME, "td") 
                                                if len(celulas) > indice_coluna_desejada_fatura:
                                                    celula_desejada_fatura = celulas[indice_coluna_desejada_fatura]
                                                    print(f"Conteúdo da coluna {indice_coluna_desejada_fatura + 1}: {celula_desejada_fatura.text}")

                                                    if palavra7.lower() in celula_desejada_fatura.text.lower() :
                                                        print(f"Pelo menos uma das palavras '{palavra7}' foi encontrada na Célula {indice_coluna_desejada_fatura + 1} da linha.")
                                                        pelo_menos_uma_palavra_encontrada_X = True
                                                        
                                            if pelo_menos_uma_palavra_encontrada_X:#FATURA ACORDO

                                                nav.switch_to.default_content()

                                                nav.find_element(By.XPATH,'//*[@id="NETAModalDialog"]/div[1]/div[1]/button/span[1]').click()#DOCUMENTOS FILTRO
                                                time.sleep(1.5)

                                                nav.switch_to.frame('ifCruscottoUtenza')
                                                iframe_1 = WebDriverWait(nav, 10).until(
                                                    EC.presence_of_element_located((By.ID,'ctl00_NetSiuCPH_lbl_crm_cruf_valstatofor'))#Status Fornecimento 
                                                )
                                                nav.find_element(By.ID,'ctl00_NetSiuCPH_rep_crm_desktop_utenza_ctl04_lb_crm_cruf_df_zoom').click()#PARCELAS
                                                time.sleep(1.5)
                                                nav.switch_to.default_content()

                                                nav.switch_to.frame('NETAModalDialogiFrame_1')
                                                iframe_1 = WebDriverWait(nav, 10).until(
                                                    EC.presence_of_element_located((By.ID,'ctl00_NetSiuCPH_ctl01_ddl_crm_bparc_filtro'))#Status Fornecimento 
                                                )

                                                nav.find_element(By.ID,'ctl00_NetSiuCPH_ctl01_ddl_crm_bparc_filtro').click()#fILTOR
                                                time.sleep(1.5)

                                                nav.find_element(By.XPATH,'//*[@id="ctl00_NetSiuCPH_ctl01_ddl_crm_bparc_filtro"]/option[2]').click()#FECHAR
                                                time.sleep(3.0)

                                                try:
                                                    tabela = WebDriverWait(nav, 10).until(
                                                        EC.presence_of_element_located((By.ID, "ctl00_NetSiuCPH_ctl01_ngv_crm_bparc"))
                                                    ) 
                                                    n = 1
                                                    while (n <= 1000):

                                                        elemento = nav.find_element(By.ID,'ctl00_upProgress')

                                                        if elemento.is_displayed():
                                                            print("O elemento está visível na página.")

                                                        else:
                                                            print("O elemento não está visível na página.")
                                                            break
                                                            
                                                        n += 1 

                                                    palavra8 = "LOTE EFETIVADO"
                                                    indice_coluna_desejada_Status = 18
                                                    pelo_menos_uma_palavra_encontrada = False

                                                    tabela_Status = nav.find_element(By.ID,'ctl00_NetSiuCPH_ctl01_ngv_crm_bparc')

                                                    linhas_status = tabela_Status.find_elements(By.TAG_NAME,"tr")
                                                    for linha in linhas_status:
                                                        # Para cada linha, extraia as células (td)
                                                        celulas = linha.find_elements(By.TAG_NAME, "td") 


                                                        if len(celulas) > indice_coluna_desejada_Status:
                                                            celula_desejada_Status = celulas[indice_coluna_desejada_Status]
                                                            print(f"Conteúdo da coluna {indice_coluna_desejada_Status + 1}: {celula_desejada_Status.text}")

                                                            if palavra8.lower() in celula_desejada_Status.text.lower() :
                                                                print(f"Pelo menos uma das palavras '{palavra8}' foi encontrada na Célula {indice_coluna_desejada_Status + 1} da linha.")
                                                                pelo_menos_uma_palavra_encontrada_Y = True
                                                                
                                                    if pelo_menos_uma_palavra_encontrada_Y:#FATURA ACORDO

                                                        def encontrar_proxima_linha_vazia_na_coluna(sheet, coluna):
                                                            linha = 1
                                                            while sheet.cell(row=linha, column=coluna).value is not None:
                                                                    linha += 1
                                                            return linha

                                                        try:
                                                            wb = openpyxl.load_workbook('Consulta.xlsx')
                                                        except FileNotFoundError:
                                                            wb = openpyxl.Workbook()

                                                        if 'Sheet' in wb.sheetnames:
                                                            ws = wb['Sheet']
                                                        else:
                                                            ws = wb.active

                                                        Hoje = datetime.now()
                                                        Data = Hoje.strftime("%d/%m/%Y")
                                                        Hora = Hoje.strftime("%H:%M:%S")    


                                                        proxima_linha_b = encontrar_proxima_linha_vazia_na_coluna(ws, coluna=4)
                                                        ws.cell(row=proxima_linha_b, column=4, value= texto )

                                                        proxima_linha_C = encontrar_proxima_linha_vazia_na_coluna(ws, coluna=5)
                                                        ws.cell(row=proxima_linha_C, column=5, value= "ACORDO EM ANDAMENTO")

                                                        proxima_linha_D = encontrar_proxima_linha_vazia_na_coluna(ws, coluna=6)
                                                        ws.cell(row=proxima_linha_D, column=6, value= Data )

                                                        proxima_linha_E = encontrar_proxima_linha_vazia_na_coluna(ws, coluna=7)
                                                        ws.cell(row=proxima_linha_E, column=7, value= Hora )
                                                                                
                                                        wb.save('Consulta.xlsx')

                                                        nav.switch_to.default_content()

                                                        nav.find_element(By.XPATH,'//*[@id="NETAModalDialog"]/div[1]/div[1]/button/span[1]').click()#DOCUMENTOS FILTRO
                                                        time.sleep(1.5)
                                                                
                                                        nav.find_element(By.XPATH,'//*[@id="ctl00_NetSiuCPH_TabCRM_bli_tcrm_cruscotti"]/li/a/span/input').click()#FECHAR
                                                        time.sleep(1.0)

                                                    else:
                                                        
                                                        palavra9 = "ACORDO ROMPIDO"
                                                        indice_coluna_desejada_Status = 18
                                                        pelo_menos_uma_palavra_encontrada = False

                                                        tabela_Status = nav.find_element(By.ID,'ctl00_NetSiuCPH_ctl01_ngv_crm_bparc')

                                                        linhas_status = tabela_Status.find_elements(By.TAG_NAME,"tr")
                                                        for linha in linhas_status:
                                                            # Para cada linha, extraia as células (td)
                                                            celulas = linha.find_elements(By.TAG_NAME, "td") 


                                                            if len(celulas) > indice_coluna_desejada_Status:
                                                                celula_desejada_Status = celulas[indice_coluna_desejada_Status]
                                                                print(f"Conteúdo da coluna {indice_coluna_desejada_Status + 1}: {celula_desejada_Status.text}")

                                                                if palavra9.lower() in celula_desejada_Status.text.lower() :
                                                                    print(f"Pelo menos uma das palavras '{palavra9}' foi encontrada na Célula {indice_coluna_desejada_Status + 1} da linha.")
                                                                    pelo_menos_uma_palavra_encontrada_Y = True
                                                                    
                                                        if pelo_menos_uma_palavra_encontrada_Y:#FATURA ACORDO

                                                            def encontrar_proxima_linha_vazia_na_coluna(sheet, coluna):
                                                                linha = 1
                                                                while sheet.cell(row=linha, column=coluna).value is not None:
                                                                        linha += 1
                                                                return linha

                                                            try:
                                                                wb = openpyxl.load_workbook('Consulta.xlsx')
                                                            except FileNotFoundError:
                                                                wb = openpyxl.Workbook()

                                                            if 'Sheet' in wb.sheetnames:
                                                                ws = wb['Sheet']
                                                            else:
                                                                ws = wb.active

                                                            Hoje = datetime.now()
                                                            Data = Hoje.strftime("%d/%m/%Y")
                                                            Hora = Hoje.strftime("%H:%M:%S")    


                                                            proxima_linha_b = encontrar_proxima_linha_vazia_na_coluna(ws, coluna=4)
                                                            ws.cell(row=proxima_linha_b, column=4, value= texto )

                                                            proxima_linha_C = encontrar_proxima_linha_vazia_na_coluna(ws, coluna=5)
                                                            ws.cell(row=proxima_linha_C, column=5, value= "DEBITO EM ABERTO - ACORDO ROMPIDO")

                                                            proxima_linha_D = encontrar_proxima_linha_vazia_na_coluna(ws, coluna=6)
                                                            ws.cell(row=proxima_linha_D, column=6, value= Data )

                                                            proxima_linha_E = encontrar_proxima_linha_vazia_na_coluna(ws, coluna=7)
                                                            ws.cell(row=proxima_linha_E, column=7, value= Hora )
                                                                                    
                                                            wb.save('Consulta.xlsx')

                                                            nav.switch_to.default_content()

                                                            nav.find_element(By.XPATH,'//*[@id="NETAModalDialog"]/div[1]/div[1]/button/span[1]').click()#DOCUMENTOS FILTRO
                                                            time.sleep(1.5)
                                                                    
                                                            nav.find_element(By.XPATH,'//*[@id="ctl00_NetSiuCPH_TabCRM_bli_tcrm_cruscotti"]/li/a/span/input').click()#FECHAR
                                                            time.sleep(1.0)                                       
                                                
                                                except TimeoutException:

                                                    def encontrar_proxima_linha_vazia_na_coluna(sheet, coluna):
                                                        linha = 1
                                                        while sheet.cell(row=linha, column=coluna).value is not None:
                                                            linha += 1
                                                        return linha

                                                    try:
                                                        wb = openpyxl.load_workbook('Consulta.xlsx')
                                                    except FileNotFoundError:
                                                        wb = openpyxl.Workbook()

                                                    if 'Sheet' in wb.sheetnames:
                                                        ws = wb['Sheet']
                                                    else:
                                                        ws = wb.active

                                                    Hoje = datetime.now()
                                                    Data = Hoje.strftime("%d/%m/%Y")
                                                    Hora = Hoje.strftime("%H:%M:%S")    

                                                    proxima_linha_b = encontrar_proxima_linha_vazia_na_coluna(ws, coluna=4)
                                                    ws.cell(row=proxima_linha_b, column=4, value= texto )

                                                    proxima_linha_C = encontrar_proxima_linha_vazia_na_coluna(ws, coluna=5)
                                                    ws.cell(row=proxima_linha_C, column=5, value= "CONTA PAGA")

                                                    proxima_linha_D = encontrar_proxima_linha_vazia_na_coluna(ws, coluna=6)
                                                    ws.cell(row=proxima_linha_D, column=6, value= Data )

                                                    proxima_linha_E = encontrar_proxima_linha_vazia_na_coluna(ws, coluna=7)
                                                    ws.cell(row=proxima_linha_E, column=7, value= Hora )
                                                                    
                                                    wb.save('Consulta.xlsx')

                                                    nav.switch_to.default_content()

                                                    nav.find_element(By.XPATH,'//*[@id="NETAModalDialog"]/div[1]/div[1]/button/span[1]').click()#DOCUMENTOS FILTRO
                                                    time.sleep(1.5)
                                                    
                                                    nav.find_element(By.XPATH,'//*[@id="ctl00_NetSiuCPH_TabCRM_bli_tcrm_cruscotti"]/li/a/span/input').click()#FECHAR
                                                    time.sleep(1.0)

                                                    break

                                            else:

                                                def encontrar_proxima_linha_vazia_na_coluna(sheet, coluna):
                                                    linha = 1
                                                    while sheet.cell(row=linha, column=coluna).value is not None:
                                                        linha += 1
                                                    return linha

                                                try:
                                                    wb = openpyxl.load_workbook('Consulta.xlsx')
                                                except FileNotFoundError:
                                                    wb = openpyxl.Workbook()

                                                if 'Sheet' in wb.sheetnames:
                                                    ws = wb['Sheet']
                                                else:
                                                    ws = wb.active

                                                Hoje = datetime.now()
                                                Data = Hoje.strftime("%d/%m/%Y")
                                                Hora = Hoje.strftime("%H:%M:%S")    

                                                proxima_linha_b = encontrar_proxima_linha_vazia_na_coluna(ws, coluna=4)
                                                ws.cell(row=proxima_linha_b, column=4, value= texto )

                                                proxima_linha_C = encontrar_proxima_linha_vazia_na_coluna(ws, coluna=5)
                                                ws.cell(row=proxima_linha_C, column=5, value= "CONTA PAGA")

                                                proxima_linha_D = encontrar_proxima_linha_vazia_na_coluna(ws, coluna=6)
                                                ws.cell(row=proxima_linha_D, column=6, value= Data )

                                                proxima_linha_E = encontrar_proxima_linha_vazia_na_coluna(ws, coluna=7)
                                                ws.cell(row=proxima_linha_E, column=7, value= Hora )
                                                                                
                                                wb.save('Consulta.xlsx')

                                                nav.switch_to.default_content()

                                                nav.find_element(By.XPATH,'//*[@id="NETAModalDialog"]/div[1]/div[1]/button/span[1]').click()#DOCUMENTOS FILTRO
                                                time.sleep(1.5)
                                                                
                                                nav.find_element(By.XPATH,'//*[@id="ctl00_NetSiuCPH_TabCRM_bli_tcrm_cruscotti"]/li/a/span/input').click()#FECHAR
                                                time.sleep(1.0)
                                                                
            except TimeoutException:

                def encontrar_proxima_linha_vazia_na_coluna(sheet, coluna):
                    linha = 1
                    while sheet.cell(row=linha, column=coluna).value is not None:
                        linha += 1
                    return linha

                try:
                    wb = openpyxl.load_workbook('Consulta.xlsx')
                except FileNotFoundError:
                    wb = openpyxl.Workbook()

                if 'Sheet' in wb.sheetnames:
                    ws = wb['Sheet']
                else:
                    ws = wb.active

                Hoje = datetime.now()
                Data = Hoje.strftime("%d/%m/%Y")
                Hora = Hoje.strftime("%H:%M:%S")    

                proxima_linha_b = encontrar_proxima_linha_vazia_na_coluna(ws, coluna=4)
                ws.cell(row=proxima_linha_b, column=4, value= texto )

                proxima_linha_C = encontrar_proxima_linha_vazia_na_coluna(ws, coluna=5)
                ws.cell(row=proxima_linha_C, column=5, value= "CONTA PAGA")

                proxima_linha_D = encontrar_proxima_linha_vazia_na_coluna(ws, coluna=6)
                ws.cell(row=proxima_linha_D, column=6, value= Data )

                proxima_linha_E = encontrar_proxima_linha_vazia_na_coluna(ws, coluna=7)
                ws.cell(row=proxima_linha_E, column=7, value= Hora )
                                                
                wb.save('Consulta.xlsx')

                nav.switch_to.default_content()

                nav.find_element(By.XPATH,'//*[@id="NETAModalDialog"]/div[1]/div[1]/button/span[1]').click()#DOCUMENTOS FILTRO
                time.sleep(1.5)
                                
                nav.find_element(By.XPATH,'//*[@id="ctl00_NetSiuCPH_TabCRM_bli_tcrm_cruscotti"]/li/a/span/input').click()#FECHAR
                time.sleep(1.0)
        else:
            print('A condição é falsa.')
            def encontrar_proxima_linha_vazia_na_coluna(sheet, coluna):
                linha = 1
                while sheet.cell(row=linha, column=coluna).value is not None:
                    linha += 1
                return linha

            try:
                wb = openpyxl.load_workbook('Consulta.xlsx')
            except FileNotFoundError:
                wb = openpyxl.Workbook()

            if 'Sheet' in wb.sheetnames:
                ws = wb['Sheet']
            else:
                ws = wb.active

            Hoje = datetime.now()
            Data = Hoje.strftime("%d/%m/%Y")
            Hora = Hoje.strftime("%H:%M:%S")    

            proxima_linha_b = encontrar_proxima_linha_vazia_na_coluna(ws, coluna=4)
            ws.cell(row=proxima_linha_b, column=4, value= texto )

            proxima_linha_C = encontrar_proxima_linha_vazia_na_coluna(ws, coluna=5)
            ws.cell(row=proxima_linha_C, column=5, value= "CICLO DE COBRANÇA ADM ENCERRADO" )

            proxima_linha_D = encontrar_proxima_linha_vazia_na_coluna(ws, coluna=6)
            ws.cell(row=proxima_linha_D, column=6, value= Data )

            proxima_linha_E = encontrar_proxima_linha_vazia_na_coluna(ws, coluna=7)
            ws.cell(row=proxima_linha_E, column=7, value= Hora )
                            
            wb.save('Consulta.xlsx')
            nav.switch_to.default_content()
            nav.find_element(By.XPATH,'//*[@id="ctl00_NetSiuCPH_TabCRM_bli_tcrm_cruscotti"]/li/a/span/input').click()#FECHAR
            time.sleep(1.0)

nav.quit()
pyautogui.alert(text='Consulta Realizada', title='', button='OK')
time.sleep(0.5)
os.startfile(R"resources\taskkill.bat")  