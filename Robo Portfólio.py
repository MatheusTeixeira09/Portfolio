from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
import time
import pyautogui
import pandas as pd
import openpyxl
from datetime import datetime
import os
from datetime import date, timedelta 

file_path_1 = 'Login.xlsx'
df = pd.read_excel(file_path_1)  

pasta_downloads = df.iloc[0, 4]
nome_arquivo_base = 'export.xls'
arquivos_na_pasta = os.listdir(pasta_downloads)
caminho_do_arquivo = os.path.join(os.path.expanduser("~"), "Downloads", "export.xls")

if nome_arquivo_base in arquivos_na_pasta:
    os.remove(caminho_do_arquivo)

else:
    time.sleep(0.5)

User = df.iloc[0, 0]
User_x = int(User)
                        
Senha = df.iloc[0, 1]
Senha_x = str(Senha)

dia_agora = date.today()
print(dia_agora)
time.sleep(0.5)
td = timedelta(120)  
x = dia_agora - td                   
Dia_120 = x.strftime("%d%m%Y")
print(Dia_120)

# Data hoje
data_atual = date.today()
data_em_texto = data_atual.strftime('%d%m%Y')
print(data_em_texto)

contrato=pyautogui.confirm(text='Escolha uma das Opções de Contratos Abaixo', title='Contrato', buttons=['Contrato 1', 'Contrato 2'])
if contrato == 'Contrato 1':

    Matricula = df.iloc[0, 2]#Contrato 1
    Matricula_x = int(Matricula)

else:
    Matricula = df.iloc[0, 3]#Contrato 2
    Matricula_x = int(Matricula)

servico = Service(ChromeDriverManager().install())
nav = webdriver.Chrome(service=servico)
nav.maximize_window()
time.sleep(1.0)

nav.get('https://nsprd.sabesp.com.br/netainf/login.aspx')#URL NET@
time.sleep(1.0)
nav.find_element(By.ID,'extended_login_Username').send_keys(User_x)#USER
time.sleep(0.5)
nav.find_element(By.ID,'extended_login_Password').send_keys(Senha_x)#SENHA
time.sleep(0.5)
nav.find_element(By.ID,'extended_login_Login').click()#BOATO LOGIN
time.sleep(0.5) 
nav.find_element(By.ID,'504').click()#CRÉDITO E COBRANÇA
time.sleep(0.5)  
nav.find_element(By.XPATH,'//*[@id="POR_153_10__504"]/div[2]').click()#CONTRATADA
time.sleep(0.5)  
nav.find_element(By.XPATH,'//*[@id="POR_153_10__2843"]/div[2]').click()#PORTFÓLIO
time.sleep(0.5)
nav.find_element(By.ID,'ctl00_NetSiuCPH_Contratada_SearchButton').click()#SETA BAIXO 
time.sleep(0.5)
nav.switch_to.frame('NETAModalDialogiFrame_1')
iframe_1 = WebDriverWait(nav, 10).until(
    EC.presence_of_element_located((By.ID, 'ctl00_NetSiuCPH_ocCodice_txtCodice'))#Seta
)
nav.find_element(By.ID,'ctl00_NetSiuCPH_ocCodice_txtCodice').send_keys(Matricula_x)#Código Contratada
time.sleep(0.5)
nav.find_element(By.ID,'ctl00_NetSiuCPH_btnCerca').click()#PESQUISAR
time.sleep(0.5)
nav.find_element(By.ID,'ctl00_NetSiuCPH_ngvLovContratada_ctl02_nbVisualizza').click()#LUPA
time.sleep(1.5)
nav.switch_to.default_content()

Data_inicio =pyautogui.confirm(text='Quer mudar a data de inicio ? \nCaso não, sera considerado os últimos 120 dias', title='Contrato', buttons=['Sim', 'Nao'])

if Data_inicio == 'Sim':

    Data_mudar = pyautogui.prompt(text='Digite a data de inicio\nPreencher somente números', title='' , default='')
    nav.find_element(By.ID,'ctl00_NetSiuCPH_txtDataAssegnazioneDA_txtIt').click()#Data inicio
    time.sleep(0.5)
    pyautogui.press('left', presses=10)
    time.sleep(0.5)
    pyautogui.write(Data_mudar)
    time.sleep(1.5)

if Data_inicio == 'Nao':

    nav.find_element(By.ID,'ctl00_NetSiuCPH_txtDataAssegnazioneDA_txtIt').click()#Data inicio
    time.sleep(0.5)
    pyautogui.press('left', presses=10)
    time.sleep(0.5)
    pyautogui.write(Dia_120)
    time.sleep(1.5)

Data_fim =pyautogui.confirm(text='Quer mudar a data final ? \nCaso não, sera considerado o dia de hoje', title='Contrato', buttons=['Sim', 'Nao'])

if Data_fim == 'Sim':

    Data_mudar_fim = pyautogui.prompt(text='Digite a data final\nPreencher somente números', title='' , default='')
    nav.find_element(By.ID,'ctl00_NetSiuCPH_txtDataAssegnazioneA_txtIt').send_keys()#Data fim
    time.sleep(0.5)
    pyautogui.press('left', presses=10)
    time.sleep(0.5)
    pyautogui.write(Data_mudar_fim)
    time.sleep(0.5)
    
if Data_fim == 'Nao':

    nav.find_element(By.ID,'ctl00_NetSiuCPH_txtDataAssegnazioneA_txtIt').send_keys(data_em_texto)#Data fim
    time.sleep(0.5)
    pyautogui.press('left', presses=10)
    time.sleep(0.5)
    dia_agora_x = date.today()
    dia_agora_str = dia_agora_x.strftime("%d%m%Y")
    pyautogui.write(dia_agora_str)
    time.sleep(0.5)

nav.find_element(By.ID,'ctl00_cphRight_btnCercaElenco').click()#Pesquisa
time.sleep(8.0)
nav.find_element(By.ID,'ctl00_NetSiuCPH_gvPortafoglio_btnmgvExport').click()#Export
time.sleep(8.0)
caminho_do_arquivo = os.path.join(os.path.expanduser("~"), "Downloads", "export.xls")
if os.path.exists(caminho_do_arquivo):
    os.startfile(caminho_do_arquivo)
else:
    time.sleep(10.0)
    os.startfile(caminho_do_arquivo)

time.sleep(5.5)
pyautogui.press('left', presses=1) 
time.sleep(0.5)
pyautogui.press('enter')
time.sleep(1.0)
pyautogui.hotkey('ctrl', 'q')

time.sleep(3.0)

file_path_1 = 'Export.xlsx'
df_2 = pd.read_excel(file_path_1)

nome_da_coluna = 'Código do Portfolio'
Pratica = df_2[nome_da_coluna].count()    
int_Pratica = int(Pratica)
print(int_Pratica)

nome_da_coluna = 'STATUS'
Status = df_2[nome_da_coluna].count() 
print(Status)

for i in range(Status, len(df_2)):

    valor_celula = df_2.iloc[i, 0]
    #valor_Res = df_2.iloc[i, 8]
    print(f"{i + 1} - {valor_celula}")
    x = (valor_celula)
    X_int = str(x)
    time.sleep(1.0)
    nav.find_element(By.ID,'ctl00_NetSiuCPH_txtCodice').send_keys(X_int)#Código
    time.sleep(0.5)
    nav.find_element(By.ID,'ctl00_cphRight_btnCercaElenco').click()#Pesquisa
    time.sleep(2.5)
    nav.find_element(By.ID,'ctl00_NetSiuCPH_gvPortafoglio_ctl02_btnFlagSelezionato').click()#Lupa
    time.sleep(2.0)
    pasta_downloads = df.iloc[0, 4]#Caminho do Downloads
    nome_arquivo_base = 'Portifolio_'
    nome_arquivo_zip = nome_arquivo_base + X_int + '.zip'
    arquivos_na_pasta = os.listdir(pasta_downloads)

    n = 1
    while (n <= 1440):
        Erro = pyautogui.locateOnScreen('resources\Erro.PNG')
        if Erro :

            nav.find_element(By.ID,'ctl00_nmp_Avvisi_Errore_btnNo').click()#Lupa
            time.sleep(2.0)
            nav.find_element(By.ID,'ctl00_NetSiuCPH_txtCodice').click()#Código
            pyautogui.press('backspace', presses=20)
            time.sleep(1.0)

            def encontrar_proxima_linha_vazia_na_coluna(sheet, coluna):
                linha = 1
                x = 0
                while sheet.cell(row=linha, column=coluna).value is not None:
                    print(linha)
                    valor_celula = df_2.iloc[x, 0]
                    print(valor_celula)
                    x +=1
                    linha += 1
                    
                return linha
            try:
                wb = openpyxl.load_workbook('Export.xlsx')
            except FileNotFoundError:
                wb = openpyxl.Workbook()

            if 'Sheet' in wb.sheetnames:
                ws = wb['Sheet']
            else:
                ws = wb.active

            proxima_linha_b = encontrar_proxima_linha_vazia_na_coluna(ws, coluna=9)

            ws.cell(row=proxima_linha_b, column=9, value="Erro")

            wb.save('Export.xlsx')


            break

        elif  nome_arquivo_zip in arquivos_na_pasta:
            print(f'O arquivo {nome_arquivo_zip} existe na pasta de downloads.')
            def encontrar_proxima_linha_vazia_na_coluna(sheet, coluna):
                linha = 1
                x = 0
                while sheet.cell(row=linha, column=coluna).value is not None:
                    print(linha)
                    valor_celula = df_2.iloc[x, 0]
                    print(valor_celula)
                    x +=1
                    linha += 1
                    
                return linha
            try:
                wb = openpyxl.load_workbook('Export.xlsx')
            except FileNotFoundError:
                wb = openpyxl.Workbook()

            if 'Sheet' in wb.sheetnames:
                ws = wb['Sheet']
            else:
                ws = wb.active

            proxima_linha_b = encontrar_proxima_linha_vazia_na_coluna(ws, coluna=9)

            ws.cell(row=proxima_linha_b, column=9, value="Baixou")

            wb.save('Export.xlsx')

            time.sleep(1.0)

            pyautogui.click(x=100, y=200)
            pyautogui.hotkey('alt', 'f4')
            time.sleep(2.0)
            nav.find_element(By.ID,'ctl00_NetSiuCPH_txtCodice').click()#Código
            pyautogui.press('backspace', presses=20)
            break
        else:
            pasta_downloads = df.iloc[0, 4]#Caminho do Downloads
            nome_arquivo_base = 'Portifolio_'
            nome_arquivo_zip = nome_arquivo_base + X_int + '.zip'
            arquivos_na_pasta = os.listdir(pasta_downloads)
            print(f'O arquivo {nome_arquivo_zip} não foi encontrado na pasta de downloads.')
            time.sleep(2.0)
            print(n)
        n += 1
    
